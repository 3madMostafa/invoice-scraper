#!/usr/bin/env python3
import json
import re
import os
import sys
from pathlib import Path
import pandas as pd
import time
import logging
from datetime import datetime, timedelta
import shutil

# Type mapping for document types
TYPE_MAPPING = {
    'i': 'Invoice',
    'c': 'Credit Note', 
    'd': 'Debit Note',
    'invoice': 'Invoice',
    'credit note': 'Credit Note',
    'debit note': 'Debit Note'
}

def setup_logging():
    """Setup logging with overwrite of old json_parser.log and proper encoding"""
    log_dir = os.path.join(os.getcwd(), "logs")
    os.makedirs(log_dir, exist_ok=True)
    
    main_log_file = os.path.join(log_dir, 'json_parser.log')
    
    if os.path.exists(main_log_file):
        os.remove(main_log_file)
    
    logger = logging.getLogger('json_parser')
    logger.setLevel(logging.INFO)
    
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    file_handler = logging.FileHandler(main_log_file, encoding='utf-8', mode='w')
    file_handler.setLevel(logging.INFO)
    
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

logger = setup_logging()

def map_document_type(type_value):
    """Map document type code to full name"""
    if not type_value:
        return "Invoice"
    
    type_lower = str(type_value).lower().strip()
    return TYPE_MAPPING.get(type_lower, str(type_value))

def load_issuer_data_from_excel():
    """Load issuer names, submission dates, and status from Excel file"""
    issuer_data_dict = {}
    
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--date', type=str)
    args = parser.parse_args()

    if args.date:
        yesterday = args.date
    else:
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%d-%m-%Y")
    excel_path = Path("logs") / f"invoices_data_{yesterday}.xlsx"
    
    if not excel_path.exists():
        logger.warning(f"Excel file not found at: {excel_path}")
        return issuer_data_dict
    
    try:
        df = pd.read_excel(excel_path)
        logger.info(f"Successfully loaded Excel file: {excel_path}")
        logger.info(f"Excel columns: {list(df.columns)}")
        
        required_columns = ['Invoice ID', 'Issuer Name', 'Submission Date']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            logger.error(f"Missing required columns in Excel: {missing_columns}")
            logger.info(f"Available columns: {list(df.columns)}")
            return issuer_data_dict
        
        for index, row in df.iterrows():
            invoice_id = str(row['Invoice ID']).strip() if pd.notna(row['Invoice ID']) else ""
            issuer_name = str(row['Issuer Name']).strip() if pd.notna(row['Issuer Name']) else ""
            submission_date = str(row['Submission Date']).strip() if pd.notna(row['Submission Date']) else ""
            status = str(row['Status']).strip() if pd.notna(row.get('Status')) else ""
            
            if invoice_id:
                issuer_data_dict[invoice_id] = {
                    'issuer_name': issuer_name,
                    'submission_date': submission_date,
                    'status': status
                }
                logger.debug(f"Loaded data: {invoice_id} → Issuer: {issuer_name}, Submission: {submission_date}, Status: {status}")
        
        logger.info(f"Loaded {len(issuer_data_dict)} records from Excel")
        
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
    
    return issuer_data_dict

def get_issuer_data_by_invoice_id(invoice_id, issuer_data_dict):
    """Get issuer name and submission date for an invoice ID from the loaded dictionary"""
    if not invoice_id or not issuer_data_dict:
        return None, None
    
    if invoice_id in issuer_data_dict:
        data = issuer_data_dict[invoice_id]
        return data.get('issuer_name'), data.get('submission_date')
    
    for key, data in issuer_data_dict.items():
        if key.lower() == invoice_id.lower():
            return data.get('issuer_name'), data.get('submission_date')
    
    return None, None

def should_exclude_supplier(issuer_name):
    """🟨 Case 2: Check if supplier should be excluded from output entirely"""
    if not issuer_name:
        return False
    
    issuer_clean = str(issuer_name).strip().replace(" ", "")
    
    excluded_suppliers = [
        "مكتبعلميامامفارما",
        "شركهثريامبي"
    ]
    
    for excluded in excluded_suppliers:
        if excluded in issuer_clean:
            logger.info(f"🟨 EXCLUDING SUPPLIER from output: {issuer_name}")
            return True
    
    return False

def is_barakat_group(issuer_name):
    """🟦 Case 3: Check if supplier is Barakat Group"""
    if not issuer_name:
        return False
    
    issuer_clean = str(issuer_name).strip().replace(" ", "").lower()
    
    barakat_patterns = [
        "بركاتجروبللتوكيلاتالتجاريةستالين",
        "بركاتجروبللتوكيلاتالتجارية",
        "بركاتجروب",
        "بركات",  # أي اسم يحتوي على "بركات"
        "barakatgroup",
        "barakat"
    ]
    
    for pattern in barakat_patterns:
        if pattern in issuer_clean:
            logger.info(f"🟦 BARAKAT GROUP detected: {issuer_name}")
            return True
    
    return False

def is_excluded_supplier(issuer_name, receiver_name=None):
    """Check if the supplier should be excluded from PO extraction (not from output)"""
    if not issuer_name:
        issuer_name = ""
    if not receiver_name:
        receiver_name = ""
    
    issuer_name_clean = str(issuer_name).strip()
    receiver_name_clean = str(receiver_name).strip()
    
    if "(شركة بي تك للتجارة والتوزيع (ش.م.م" in issuer_name_clean:
        return True, "لا يتم الربط مع المورد بي تك"
    
    if "الكرنك لتجارة السيارات" in issuer_name_clean:
        return True, "لا يتم الربط مع شركه الكرنك"
    
    if "مانترا للسيارات" in issuer_name_clean:
        return True, ""
    
    if "شركة لؤلؤه البحرين لخدمه السيارات المرسيدس عبد الرازق وشركاه" in issuer_name_clean:
        return True, "لا يتم الربط مع المورد لؤلؤه البحرين"
    
    if "شركة لؤلؤه البحرين لخدمه السيارات المرسيدس عبد الرازق وشركاه" in receiver_name_clean:
        return True, "لا يتم الربط مع المورد لؤلؤه البحرين"
    
    return False, ""

def convert_arabic_numbers_to_english(text):
    """Convert Arabic-Indic numerals to English numerals"""
    arabic_to_english = {
        '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
        '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9'
    }
    
    for arabic, english in arabic_to_english.items():
        text = text.replace(arabic, english)
    
    return text

def find_numbers_4_to_6_digits(text, source_field="unknown"):
    """Extract all numbers with 4-6 digits, applying comprehensive PO number filtering rules"""
    if not text:
        return []
    
    text = str(text)
    text = convert_arabic_numbers_to_english(text)
    
    # CRITICAL: Skip any sequence of 7+ consecutive digits
    long_number_pattern = r'\d{7,}'
    if re.search(long_number_pattern, text):
        long_numbers = re.findall(long_number_pattern, text)
        logger.info(f"[{source_field}] Found numbers with 7+ digits (will be skipped): {long_numbers}")
        for long_num in long_numbers:
            text = text.replace(long_num, ' ')
    
    po_keywords = [
        'po', 'po num', 'ponum', 'po/', 'po number', 'po reference',
        'purchase order', 'purchaseorder', 'purchaseorderreference', 
        'purchaseorderdescription', 'salesorderreference',
        'order number', 'ordernumber',
        'كود', 'اوردر رقم', 'اوردررقم', 'اوردر', 'رقماوردر',  # إضافة النسخ الملزقة
        'حجز رقم', 'حجزرقم',
        'وجبات', 'وجباتغذائية', 'وجبات غذائية',
        'موافقة', 'الموافقة', 'طلب', 'طلبات', 'الطلب', 'موافقات', 'الموافقات',
        'أرقامالموافقات', 'أرقام الموافقات', 'ارقام الموافقات',
        'po no', 'pono', 'p.o', 'p.o.', 'po.', 'رقم po', 'رقمpo'
    ]
    
    avoid_keywords = [
        'ltr', 'لتر', 'kg', 'kgm', 'egp', 'usd', 'eur', 'gbp',
        'وات', 'watt', 'w', 'cm', 'mm', 'meter', 'متر',
        'gram', 'جرام', 'piece', 'قطعة', 'box', 'صندوق',
        'postal', 'code', 'zip', 'بريد', 'بريدي', 'كود بريدي',
        'chassis', 'شاسيه', 'لوحة', 'plate', 'vin', 'سنة صنع',
        'year', 'delivery', 'دليفري', 'توصيل'
    ]
    
    financial_keywords = [
        'بقيمة', 'قيمة', 'جنيه', 'جنية', 'ج', 'جم', 'جنيهات',
        'value', 'worth', 'cost', 'price', 'amount'
    ]
    
    # توسيع قائمة السنوات المستبعدة
    avoid_years = ['2023', '2024', '2025', '2026', '2027', '2028', '2029', '2030']
    
    # أرقام بريدية شائعة يجب استبعادها
    common_postal_codes = ['1111', '11111', '12345', '54321', '99999']
    
    # PRIORITY 1: "(PONUM174928/174261)" pattern
    logger.info(f"[{source_field}] Checking for (PONUM...) pattern")
    parentheses_po_pattern = r'\(ponum([^\)]+)\)'
    parentheses_matches = re.findall(parentheses_po_pattern, text, re.IGNORECASE)
    
    if parentheses_matches:
        logger.info(f"[{source_field}] FOUND (PONUM...) pattern! Matches: {parentheses_matches}")
        all_parentheses_po = []
        for number_string in parentheses_matches:
            numbers = re.findall(r'\d{4,6}', number_string)
            logger.info(f"[{source_field}] Captured: '{number_string}', Extracted: {numbers}")
            
            for number in numbers:
                if (not number.startswith('0') and 
                    number not in avoid_years and 
                    number not in common_postal_codes and
                    4 <= len(number) <= 6 and
                    not number.endswith('000') and  # فقط استبعاد إذا انتهى بـ 00 (مثل 1100, 2300)
                    not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999'))):
                    all_parentheses_po.append(number)
                    logger.info(f"[{source_field}] Valid PO: {number}")
        
        if all_parentheses_po:
            logger.info(f"[{source_field}] RETURNING {len(all_parentheses_po)} numbers from (PONUM...): {all_parentheses_po}")
            return all_parentheses_po
    
    # PRIORITY 2: "PONUMFORDr/AntonyTharwatis8389" pattern
    logger.info(f"[{source_field}] Checking for PONUM FOR...is pattern")
    multiline_po_pattern = r'ponumfor[^\d]*?(?:is|:)\s*(\d{4,6})'
    multiline_matches = re.findall(multiline_po_pattern, text, re.IGNORECASE)
    
    if multiline_matches:
        logger.info(f"[{source_field}] FOUND PONUM FOR...is pattern! Matches: {multiline_matches}")
        all_multiline_po = []
        for number in multiline_matches:
            if (not number.startswith('0') and 
                number not in avoid_years and 
                number not in common_postal_codes and
                4 <= len(number) <= 6 and
                not number.endswith('00') and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999'))):
                all_multiline_po.append(number)
                logger.info(f"[{source_field}] Valid PO from 'FOR...is': {number}")
        
        if all_multiline_po:
            logger.info(f"[{source_field}] RETURNING {len(all_multiline_po)} numbers from PONUM FOR: {all_multiline_po}")
            return all_multiline_po
    
    # PRIORITY 3: "P08561" format (PO stuck to number)
    po_stuck_pattern = r'(?i)p0?(\d{4,6})\b'
    po_stuck_matches = re.findall(po_stuck_pattern, text)
    
    if po_stuck_matches:
        stuck_po_numbers = []
        for number in po_stuck_matches:
            if (not number.startswith('0') and 
                number not in avoid_years and
                number not in common_postal_codes and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                not number.endswith('00')):
                stuck_po_numbers.append(number)
                logger.info(f"[{source_field}] Found stuck PO format: P0{number} → extracted: {number}")
        
        if stuck_po_numbers:
            return stuck_po_numbers
        
    # PRIORITY 3.5: "P.o176717" format (P.o stuck to number with dot)
    po_dot_stuck_pattern = r'(?i)p\.o?(\d{4,6})\b'
    po_dot_stuck_matches = re.findall(po_dot_stuck_pattern, text)

    if po_dot_stuck_matches:
        dot_stuck_po_numbers = []
        for number in po_dot_stuck_matches:
            if (not number.startswith('0') and 
                number not in avoid_years and
                number not in common_postal_codes and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                not number.endswith('00')):
                dot_stuck_po_numbers.append(number)
                logger.info(f"[{source_field}] Found stuck P.o format: P.o{number} → extracted: {number}")
        
        if dot_stuck_po_numbers:
            return dot_stuck_po_numbers

    # PRIORITY 3.6: "PONO:-406" format (allow 3-digit numbers ONLY with PONO keyword)
    pono_short_pattern = r'(?i)po\s*no\s*[:\-]\s*(\d{3,6})'
    pono_short_matches = re.findall(pono_short_pattern, text)

    if pono_short_matches:
        pono_short_numbers = []
        for number in pono_short_matches:
            # Allow 3-digit numbers ONLY for PONO pattern
            if (not number.startswith('0') and 
                number not in avoid_years and
                number not in common_postal_codes and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999'))):
                pono_short_numbers.append(number)
                logger.info(f"[{source_field}] Found PONO short format: PONO:{number}")
        
        if pono_short_numbers:
            return pono_short_numbers
    
    # PRIORITY 4: PO numbers with prefixes
    po_prefix_pattern = r'(?i)(po[#\-_\s]*[a-z]*\s*)(\d{4,6})'
    po_prefix_matches = re.findall(po_prefix_pattern, text)
    
    if po_prefix_matches:
        prefixed_po_numbers = []
        for prefix, number in po_prefix_matches:
            if (not number.startswith('0') and 
                number not in avoid_years and
                number not in common_postal_codes and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                not number.endswith('00')):
                prefixed_po_numbers.append(number)
                logger.info(f"[{source_field}] Found PO with prefix: {prefix.strip()}{number} → extracted: {number}")
        
        if prefixed_po_numbers:
            return prefixed_po_numbers
    
    # PRIORITY 5: Multiple PO numbers with delimiters (including - and +)
    multi_po_pattern = r'(?i)(po\s*/?(?:num|number|no)?|purchase\s*order|order\s*number|ordernumber|كود|اوردر\s*رقم|اوردررقم|اوردر|رقماوردر|حجز\s*رقم|حجزرقم|وجبات(?:غذائية)?|موافقة|الموافقة|طلب|طلبات|الطلب|موافقات|الموافقات|أرقامالموافقات|أرقام\s*الموافقات|ارقام\s*الموافقات|p\.?o\.?(?:no)?)\s*[:/\-]?\s*(\d{4,6}(?:\s*[/\-\\,،\s+]+\s*\d{4,6})*)'
    multi_po_matches = re.findall(multi_po_pattern, text)
    
    if multi_po_matches:
        all_po_numbers = []
        for keyword, number_string in multi_po_matches:
            numbers = re.findall(r'\d{4,6}', number_string)
            logger.info(f"[{source_field}] Extracted numbers from '{number_string}': {numbers}")
            
            valid_numbers = []
            for number in numbers:
                if (not number.startswith('0') and 
                    number not in avoid_years and
                    number not in common_postal_codes and
                    not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                    4 <= len(number) <= 6 and
                    not number.endswith('00')):
                    valid_numbers.append(number)
            
            if valid_numbers:
                logger.info(f"[{source_field}] Found multiple PO numbers: {valid_numbers}")
                all_po_numbers.extend(valid_numbers)
        
        if all_po_numbers:
            unique_po_numbers = []
            seen = set()
            for num in all_po_numbers:
                if num not in seen:
                    unique_po_numbers.append(num)
                    seen.add(num)
            return unique_po_numbers
    
    # PRIORITY 6: "(PO NO.169741)" pattern
    po_parentheses_pattern = r'\((?:po\s*no\.?|po\s*num\.?|po)\s*(\d{4,6})\)'
    po_parentheses_matches = re.findall(po_parentheses_pattern, text, re.IGNORECASE)
    
    if po_parentheses_matches:
        logger.info(f"[{source_field}] Found (PO NO...) pattern! Matches: {po_parentheses_matches}")
        valid_po_paren = []
        for number in po_parentheses_matches:
            if (not number.startswith('0') and 
                number not in avoid_years and
                number not in common_postal_codes and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                4 <= len(number) <= 6 and
                not number.endswith('00')):
                valid_po_paren.append(number)
                logger.info(f"[{source_field}] Valid PO from (PO NO...): {number}")
        
        if valid_po_paren:
            return valid_po_paren
    
    # PRIORITY 7: PO numbers in simple parentheses near PO keywords
    parentheses_pattern = r'\((\d{4,6})[A-Za-z]?\)'
    parentheses_matches = re.findall(parentheses_pattern, text)
    
    if parentheses_matches:
        valid_parentheses_numbers = []
        for number in parentheses_matches:
            if (not number.startswith('0') and 
                number not in avoid_years and
                number not in common_postal_codes and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                4 <= len(number) <= 6 and
                not number.endswith('00')):
                number_pos = text.find(f'({number}')
                if number_pos != -1:
                    before_text = text[max(0, number_pos-50):number_pos].lower()
                    has_po_keyword = any(keyword in before_text for keyword in po_keywords)
                    if has_po_keyword:
                        valid_parentheses_numbers.append(number)
                        logger.info(f"[{source_field}] Found PO number in parentheses: {number}")
        
        if valid_parentheses_numbers:
            return valid_parentheses_numbers
    
    # PRIORITY 8: Single PO numbers with keywords
    po_pattern = r'(?i)(po\s*/?(?:num)?|purchase\s*order|order\s*number|ordernumber|كود|اوردر\s*رقم|اوردررقم|اوردر|رقماوردر|حجز\s*رقم|حجزرقم|وجبات|موافقة|الموافقة|طلب|طلبات|الطلب|موافقات|الموافقات|أرقامالموافقات|أرقام\s*الموافقات|ارقام\s*الموافقات|po\s*no|pono|p\.?o\.?)\s*[:/\-]?\s*(\d{4,6})'
    po_matches = re.findall(po_pattern, text)
    
    if po_matches:
        po_numbers = []
        for keyword, number in po_matches:
            if (not number.startswith('0') and 
                number not in avoid_years and
                number not in common_postal_codes and
                not number.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                not number.endswith('00')):
                po_numbers.append(number)
                logger.info(f"[{source_field}] Found PO number: {number} (keyword: {keyword})")
        
        if po_numbers:
            return po_numbers
    
    # PRIORITY 9: General 4-6 digit numbers with comprehensive filtering
    pattern = r'(\d{4,6})'
    matches = re.findall(pattern, text)
    
    valid_numbers = []
    
    for match in matches:
        if not (4 <= len(match) <= 6):
            continue
        
        # استبعاد الأرقام البريدية الشائعة
        if match in common_postal_codes:
            logger.info(f"[{source_field}] Skipping common postal code: {match}")
            continue
        
        # استبعاد الأرقام التي تبدأ بنمط متكرر
        if match.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')):
            logger.info(f"[{source_field}] Skipping repetitive pattern: {match}")
            continue
            
        if match.startswith('0'):
            logger.info(f"[{source_field}] Skipping number starting with 0: {match}")
            continue
        
        # استبعاد الأرقام التي تنتهي بـ 00 فقط (مثل 1100، 2300) - ولكن السماح بـ 00 في الوسط (مثل 177002)
        if match.endswith('000'):
            logger.info(f"[{source_field}] Skipping number ending with 00: {match}")
            continue
        
        match_pos = text.find(match)
        if match_pos == -1:
            continue
        
        # Check for vehicle-related context
        before_text_vehicle = text[max(0, match_pos-100):match_pos].lower()
        after_text_vehicle = text[match_pos + len(match):match_pos + len(match) + 100].lower()
        vehicle_context = before_text_vehicle + after_text_vehicle
        
        vehicle_keywords = ['chassis', 'شاسيه', 'لوحة', 'plate', 'vin', 'رقم لوحة', 'رقم شاسيه']
        if any(keyword in vehicle_context for keyword in vehicle_keywords):
            logger.info(f"[{source_field}] Skipping vehicle-related number: {match}")
            continue
        
        # Check if it's a postal code (in address context)
        postal_keywords = ['postal', 'code', 'zip', 'بريد', 'بريدي', 'كود بريدي', 'buildingNumber', 'postalCode']
        if source_field in ['receiver.address.buildingNumber', 'receiver.address.postalCode'] or \
           any(keyword in vehicle_context for keyword in postal_keywords):
            logger.info(f"[{source_field}] Skipping postal code: {match}")
            continue
        
        # Check if it's in a delivery/product description without PO keyword
        delivery_keywords = ['delivery', 'دليفري', 'توصيل', 'description', 'وصف']
        has_delivery_context = any(keyword in vehicle_context for keyword in delivery_keywords)
        has_po_keyword_nearby = any(keyword in vehicle_context for keyword in po_keywords)
        
        if has_delivery_context and not has_po_keyword_nearby:
            logger.info(f"[{source_field}] Skipping delivery/product description number: {match}")
            continue
        
        dimension_patterns = [
            rf'\b\d{{1,4}}[/x×]\d{{1,4}}[/x×]{re.escape(match)}\b',
            rf'\b{re.escape(match)}[/x×]\d{{1,4}}[/x×]\d{{1,4}}\b',
            rf'\b\d{{1,4}}[/x×]{re.escape(match)}[/x×]\d{{1,4}}\b',
            rf'\b\d{{1,4}}\s*[x×]\s*{re.escape(match)}\b',
            rf'\b{re.escape(match)}\s*[x×]\s*\d{{1,4}}\b',
            rf'\b{re.escape(match)}\s*[x×]\s*\d{{4,6}}\b',
            rf'\b\d{{4,6}}\s*[x×]\s*{re.escape(match)}\b'
        ]
        
        is_dimension = False
        for pattern in dimension_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                logger.info(f"[{source_field}] Skipping dimension: {match}")
                is_dimension = True
                break
        
        if is_dimension:
            continue
        
        start_check = max(0, match_pos - 10)
        end_check = min(len(text), match_pos + len(match) + 10)
        surrounding = text[start_check:end_check]
        
        model_patterns = [
            rf'[A-Za-z]+{re.escape(match)}[A-Za-z]*',
            rf'[A-Za-z]*{re.escape(match)}[A-Za-z]+',
            rf'[A-Za-z]+{re.escape(match)}',
            rf'{re.escape(match)}[A-Za-z]+',
        ]
        
        is_model_number = False
        for pattern in model_patterns:
            if re.search(pattern, surrounding):
                before_text = text[max(0, match_pos-50):match_pos].lower()
                has_nearby_po_keyword = any(keyword in before_text for keyword in po_keywords)
                
                if not has_nearby_po_keyword:
                    logger.info(f"[{source_field}] Skipping model number: {match}")
                    is_model_number = True
                    break
        
        if is_model_number:
            continue
            
        before_text = text[max(0, match_pos-40):match_pos].lower()
        after_text = text[match_pos + len(match):match_pos + len(match) + 40].lower()
        context = before_text + after_text
        
        is_financial_value = False
        in_parentheses = f'({match}' in text or f'({match}B)' in text or f'({match}G)' in text
        
        near_po_keyword = False
        for po_keyword in po_keywords:
            po_keyword_pattern = rf'(?i){re.escape(po_keyword)}.{{0,30}}{re.escape(match)}'
            if re.search(po_keyword_pattern, text):
                near_po_keyword = True
                break
        
        if not in_parentheses and not near_po_keyword:
            for financial_word in financial_keywords:
                financial_pattern = rf'(?i){re.escape(financial_word)}\s*{re.escape(match)}'
                if re.search(financial_pattern, text):
                    logger.info(f"[{source_field}] Skipping financial value: {match}")
                    is_financial_value = True
                    break
        
        if is_financial_value:
            continue
        
        near_units = any(unit in context for unit in avoid_keywords)
        if near_units:
            logger.info(f"[{source_field}] Skipping unit-related number: {match}")
            continue
            
        if match in avoid_years:
            logger.info(f"[{source_field}] Skipping year: {match}")
            continue
        
        # استبعاد سنوات الصنع (4 digits starting with 1 or 2)
        if len(match) == 4 and match[0] in ['1', '2']:
            year_keywords = ['year', 'سنة', 'صنع', 'موديل', 'model']
            if any(keyword in context for keyword in year_keywords):
                logger.info(f"[{source_field}] Skipping manufacturing year: {match}")
                continue
            
        if match.endswith('000') or match.startswith('999'):
            logger.info(f"[{source_field}] Skipping common pattern: {match}")
            continue
        
        price_pattern = rf'\b{re.escape(match)}\.?\d{{0,2}}\s*(egp|usd|eur|gbp|جنيه|جنية|دولار|ج|جم)\b'
        if re.search(price_pattern, context, re.IGNORECASE):
            logger.info(f"[{source_field}] Skipping price: {match}")
            continue
        
        if match.startswith(('0100', '0101', '0102', '0106', '0109', '0110', '0111', '0112', '0114', '0115', '0120', '0121', '0122', '0127', '0128', '0150', '0151', '0152', '0155', '0156')):
            logger.info(f"[{source_field}] Skipping phone number: {match}")
            continue
            
        logger.info(f"[{source_field}] Valid PO candidate: {match}")
        valid_numbers.append(match)
    
    return valid_numbers

def extract_po_from_json_fields(document_data):
    """🟦 Case 3: Extract PO directly from JSON fields for Barakat Group"""
    po_numbers = []
    
    if 'salesOrderReference' in document_data and document_data['salesOrderReference']:
        sales_ref = str(document_data['salesOrderReference']).strip()
        if sales_ref:
            logger.info(f"🟦 BARAKAT: Found salesOrderReference = {sales_ref}")
            po_numbers.append(sales_ref)
    
    if 'purchaseOrderReference' in document_data and document_data['purchaseOrderReference']:
        purchase_ref = str(document_data['purchaseOrderReference']).strip()
        if purchase_ref:
            logger.info(f"🟦 BARAKAT: Found purchaseOrderReference = {purchase_ref}")
            po_numbers.append(purchase_ref)
    
    # 🟦 BARAKAT: ابحث في receiver.name (اسم الشركة المرسل إليها)
    if 'receiver' in document_data and isinstance(document_data['receiver'], dict):
        receiver_name = document_data['receiver'].get('name', '')
        if receiver_name:
            receiver_name_str = str(receiver_name)
            logger.info(f"🟦 BARAKAT: Checking receiver.name = {receiver_name_str[:100]}")
            numbers = find_numbers_4_to_6_digits(receiver_name_str, "receiver.name [BARAKAT]")
            if numbers:
                logger.info(f"🟦 BARAKAT: Found PO in receiver.name: {numbers}")
                po_numbers.extend(numbers)
        
        # 🟦 BARAKAT: ابحث في receiver.address.landmark و additionalInformation
        if 'address' in document_data['receiver'] and isinstance(document_data['receiver']['address'], dict):
            address = document_data['receiver']['address']
            
            if 'landmark' in address and address['landmark']:
                landmark = str(address['landmark'])
                logger.info(f"🟦 BARAKAT: Checking receiver.address.landmark = {landmark}")
                numbers = find_numbers_4_to_6_digits(landmark, "receiver.address.landmark [BARAKAT]")
                if numbers:
                    logger.info(f"🟦 BARAKAT: Found PO in landmark: {numbers}")
                    po_numbers.extend(numbers)
            
            if 'additionalInformation' in address and address['additionalInformation']:
                additional_info = str(address['additionalInformation'])
                logger.info(f"🟦 BARAKAT: Checking receiver.address.additionalInformation = {additional_info}")
                numbers = find_numbers_4_to_6_digits(additional_info, "receiver.address.additionalInformation [BARAKAT]")
                if numbers:
                    logger.info(f"🟦 BARAKAT: Found PO in additionalInformation: {numbers}")
                    po_numbers.extend(numbers)
    
    # 🟦 BARAKAT: ابحث في invoiceLines أيضاً
    if 'invoiceLines' in document_data and isinstance(document_data['invoiceLines'], list):
        logger.info(f"🟦 BARAKAT: Checking {len(document_data['invoiceLines'])} invoice lines")
        for i, line in enumerate(document_data['invoiceLines']):
            if isinstance(line, dict) and 'description' in line and line['description']:
                desc = str(line['description'])
                logger.info(f"🟦 BARAKAT: Checking invoice line {i+1}: {desc[:100]}")
                numbers = find_numbers_4_to_6_digits(desc, f"invoiceLines[{i}].description [BARAKAT]")
                if numbers:
                    logger.info(f"🟦 BARAKAT: Found PO in invoice line {i+1}: {numbers}")
                    po_numbers.extend(numbers)
    
    # 🟦 BARAKAT: ابحث قبل رقم التسجيل مباشرة
    if 'receiver' in document_data and isinstance(document_data['receiver'], dict):
        receiver_id = document_data['receiver'].get('id', '')
        if receiver_id:
            full_text = json.dumps(document_data, ensure_ascii=False)
            receiver_id_pattern = rf'(\d{{4,6}})\s*[،,\-\s]*{re.escape(str(receiver_id))}'
            matches = re.findall(receiver_id_pattern, full_text)
            
            if matches:
                logger.info(f"🟦 BARAKAT: Found PO before receiver registration {receiver_id}: {matches}")
                for match in matches:
                    if (not match.startswith('0') and 
                        match not in ['1111', '11111', '12345', '54321', '99999'] and
                        not match.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                        not match.endswith('00')):
                        po_numbers.append(match)
                        logger.info(f"🟦 BARAKAT: Valid PO before receiver registration: {match}")
    
    # إزالة التكرارات
    unique_po = []
    seen = set()
    for num in po_numbers:
        if num not in seen:
            unique_po.append(num)
            seen.add(num)
    
    return unique_po

def extract_po_from_areas(document_data):
    """Extract PO number from purchaseOrderReference, salesOrderReference, proformaInvoiceNumber, invoiceLines, receiver.name, and address fields"""
    all_po_numbers = []
    
    # إضافة proformaInvoiceNumber
    if 'proformaInvoiceNumber' in document_data:
        proforma_ref = document_data['proformaInvoiceNumber']
        if proforma_ref:
            numbers = find_numbers_4_to_6_digits(str(proforma_ref), "proformaInvoiceNumber")
            if numbers:
                logger.info(f"Found in proformaInvoiceNumber: {numbers}")
                all_po_numbers.extend(numbers)
    
    if 'purchaseOrderReference' in document_data:
        po_ref = document_data['purchaseOrderReference']
        if po_ref:
            numbers = find_numbers_4_to_6_digits(str(po_ref), "purchaseOrderReference")
            if numbers:
                logger.info(f"Found in purchaseOrderReference: {numbers}")
                all_po_numbers.extend(numbers)
    
    if 'salesOrderReference' in document_data:
        sales_ref = document_data['salesOrderReference']
        if sales_ref:
            numbers = find_numbers_4_to_6_digits(str(sales_ref), "salesOrderReference")
            if numbers:
                logger.info(f"Found in salesOrderReference: {numbers}")
                all_po_numbers.extend(numbers)
    
    # البحث عن PO قبل رقم التسجيل (للموردين مثل بركات)
    if 'issuer' in document_data and isinstance(document_data['issuer'], dict):
        issuer_id = document_data['issuer'].get('id', '')
        if issuer_id:
            # البحث عن نص كامل للفاتورة
            full_text = json.dumps(document_data, ensure_ascii=False)
            
            # البحث عن الأرقام قبل رقم التسجيل
            issuer_id_pattern = rf'(\d{{4,6}})\s*[،,\-\s]*{re.escape(str(issuer_id))}'
            matches = re.findall(issuer_id_pattern, full_text)
            
            if matches:
                logger.info(f"Found PO number(s) before registration number {issuer_id}: {matches}")
                for match in matches:
                    if (not match.startswith('0') and 
                        match not in ['1111', '11111', '12345', '54321', '99999'] and
                        not match.startswith(('111', '222', '333', '444', '555', '666', '777', '888', '999')) and
                        not match.endswith('00')):
                        all_po_numbers.append(match)
                        logger.info(f"Valid PO before registration: {match}")
    
    if 'receiver' in document_data and isinstance(document_data['receiver'], dict):
        # البحث في receiver.address.landmark (هنا غالباً تكتب PO)
        if 'address' in document_data['receiver'] and isinstance(document_data['receiver']['address'], dict):
            address = document_data['receiver']['address']
            
            # 🔍 البحث في landmark
            if 'landmark' in address and address['landmark']:
                landmark = str(address['landmark'])
                logger.info(f"Checking receiver.address.landmark: {landmark}")
                numbers = find_numbers_4_to_6_digits(landmark, "receiver.address.landmark")
                if numbers:
                    logger.info(f"Found in receiver.address.landmark: {numbers}")
                    all_po_numbers.extend(numbers)
            
            # 🔍 البحث في additionalInformation أيضاً
            if 'additionalInformation' in address and address['additionalInformation']:
                additional_info = str(address['additionalInformation'])
                logger.info(f"Checking receiver.address.additionalInformation: {additional_info}")
                numbers = find_numbers_4_to_6_digits(additional_info, "receiver.address.additionalInformation")
                if numbers:
                    logger.info(f"Found in receiver.address.additionalInformation: {numbers}")
                    all_po_numbers.extend(numbers)
        
        # تخطي buildingNumber و postalCode تماماً (غالباً أرقام بريدية)
        logger.info("Skipping receiver.address fields (buildingNumber, postalCode) - typically postal codes")
        
        receiver_name = document_data['receiver'].get('name', '')
        if receiver_name:
            receiver_name_str = str(receiver_name)
            po_keywords_check = ['po', 'purchase', 'order', 'كود', 'اوردر', 'حجز', 'وجبات', 'موافقة', 'طلب']
            has_po_keyword = any(keyword in receiver_name_str.lower() for keyword in po_keywords_check)
            
            if has_po_keyword:
                numbers = find_numbers_4_to_6_digits(receiver_name_str, "receiver.name")
                if numbers:
                    logger.info(f"Found in receiver.name: {numbers}")
                    all_po_numbers.extend(numbers)
            else:
                logger.info(f"Skipping receiver.name (no PO keywords found): {receiver_name_str[:50]}")
    
    if 'invoiceLines' in document_data and isinstance(document_data['invoiceLines'], list):
        logger.info(f"Checking {len(document_data['invoiceLines'])} invoice lines")
        for i, line in enumerate(document_data['invoiceLines']):
            if isinstance(line, dict) and 'description' in line and line['description']:
                desc = str(line['description'])
                desc_clean = desc.lower().replace(" ", "").replace("\n", "").replace("\r", "")
                
                # التحقق من وجود كلمة PO في الوصف
                po_keywords_check = [
                    'po', 'p.o', 'purchase', 'order', 'ordernumber',
                    'كود', 'اوردر', 'اوردررقم', 'رقماوردر',
                    'موافقة', 'طلب', 'وجبات', 'وجباتغذائية', 'حجز'
                ]
                has_po_keyword = any(keyword in desc_clean for keyword in po_keywords_check)
                
                # التحقق من وجود رقم في بداية السطر أو بعد سطر جديد
                has_leading_number = bool(re.search(r'(^|\n|\r)\s*\d{4,6}', desc))
                
                if has_po_keyword or has_leading_number:
                    numbers = find_numbers_4_to_6_digits(desc, f"invoiceLines[{i}].description")
                    if numbers:
                        logger.info(f"Found in invoice line {i+1}: {numbers}")
                        all_po_numbers.extend(numbers)
                else:
                    logger.info(f"Skipping invoice line {i+1} (no PO keywords or leading number): {desc[:50]}")
    
    unique_numbers = []
    seen = set()
    for num in all_po_numbers:
        if num not in seen:
            unique_numbers.append(num)
            seen.add(num)
    
    return unique_numbers

def get_issuer_name_from_json(document_data):
    """🟩 Case 1: Get issuer name from JSON document.issuer.name"""
    try:
        if 'issuer' in document_data and isinstance(document_data['issuer'], dict):
            issuer_name = document_data['issuer'].get('name', '')
            if issuer_name:
                logger.info(f"🟩 Found issuer name in JSON: {issuer_name}")
                return str(issuer_name).strip()
    except Exception as e:
        logger.error(f"Error extracting issuer name from JSON: {e}")
    
    return ""

def is_cancelled_or_rejected_invoice(document_data, excel_status=""):
    """Check if invoice is cancelled or rejected from JSON or Excel status"""
    
    # Check Excel status first
    if excel_status:
        excel_status_lower = str(excel_status).lower()
        if 'cancel' in excel_status_lower or 'ملغ' in excel_status_lower:
            return True, "Cancelled"
        if 'reject' in excel_status_lower or 'مرفوض' in excel_status_lower:
            return True, "Rejected"
    
    # Check JSON status field
    try:
        status = document_data.get('status', '').lower()
        if 'cancel' in status or 'ملغ' in status:
            return True, "Cancelled"
        if 'reject' in status or 'مرفوض' in status or 'invalid' in status:
            return True, "Rejected"
    except:
        pass
    
    return False, None

def process_json_file(file_path, issuer_data_dict):
    """Process a single JSON file and extract required data"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        logger.info(f"Processing file: {file_path.name}")
        
        document_data = data
        if 'document' in data and isinstance(data['document'], str):
            try:
                document_data = json.loads(data['document'])
                logger.info("Found nested JSON in 'document' field")
            except json.JSONDecodeError:
                logger.error("Could not parse nested JSON in 'document' field")
        
        original_from = data.get("issuerName", "")
        invoice_id = data.get("uuid", "") or data.get("internalId", "")
        
        # Get data from Excel
        excel_issuer_name, excel_submission_date = get_issuer_data_by_invoice_id(invoice_id, issuer_data_dict)
        
        # Get Excel status
        excel_status = ""
        if invoice_id in issuer_data_dict:
            excel_status = issuer_data_dict[invoice_id].get('status', '')
        
        # Handle issuer name
        if excel_issuer_name and excel_issuer_name.strip() and excel_issuer_name.strip() != "غير محدد":
            final_from = ' '.join(excel_issuer_name.split())  # حذف المسافات الزائدة
            logger.info(f"Using issuer name from Excel: {invoice_id} → {excel_issuer_name}")
        else:
            json_issuer_name = get_issuer_name_from_json(document_data)
            if json_issuer_name:
                final_from = ' '.join(json_issuer_name.split())  # حذف المسافات الزائدة
                logger.info(f"🟩 Using issuer name from JSON (Excel was '{excel_issuer_name}'): {json_issuer_name}")
            else:
                final_from = ' '.join(original_from.split())  # حذف المسافات الزائدة
                logger.info(f"Using original issuer name: {original_from}")
        
        # Handle submission date
        if excel_submission_date and excel_submission_date.strip() and excel_submission_date.strip() != "غير محدد":
            final_submission_date = excel_submission_date
            logger.info(f"Using submission date from Excel: {invoice_id} → {excel_submission_date}")
        else:
            final_submission_date = ""
            logger.info(f"No valid submission date found in Excel for: {invoice_id}")
        
        receiver_name = ""
        if 'receiver' in document_data and isinstance(document_data['receiver'], dict):
            receiver_name = document_data['receiver'].get('name', '')
        
        raw_type = data.get("typeName", "Invoice")
        mapped_type = map_document_type(raw_type)
        logger.info(f"Type mapping: '{raw_type}' → '{mapped_type}'")
        
        # 🆕 CHECK IF CANCELLED OR REJECTED
        is_cancelled, cancel_status = is_cancelled_or_rejected_invoice(document_data, excel_status)

        if is_cancelled:
            logger.info(f"⚠️ Invoice is {cancel_status} - will add note in PO field")
            po_number = f"{cancel_status}"
        elif mapped_type == "Credit Note":
            logger.info("Document type is Credit Note - skipping PO extraction")
            po_number = "لا يتم الربط مع الفواتير ال credit note"
        elif should_exclude_supplier(final_from):  # 🟨 إضافة الشرط هنا
            logger.info(f"🟨 Excluded supplier - adding note: {final_from}")
            po_number = f"لا يتم الربط مع {final_from}"
        else:
            # Normal PO extraction
            if is_barakat_group(final_from):
                po_numbers = extract_po_from_json_fields(document_data)
                if po_numbers:
                    po_number = ", ".join(po_numbers)
                    logger.info(f"🟦 BARAKAT GROUP - PO from JSON fields: {po_number}")
                else:
                    po_number = ""
                    logger.warning(f"🟦 BARAKAT GROUP - No PO found in JSON fields")
            else:
                is_excluded, exclusion_message = is_excluded_supplier(final_from, receiver_name)
                
                if is_excluded:
                    if exclusion_message:
                        logger.info(f"Excluding supplier from PO extraction: {final_from}")
                        logger.info(f"Exclusion message: {exclusion_message}")
                        po_number = exclusion_message
                    else:
                        logger.info(f"Skipping PO extraction for issuer: {final_from}")
                        po_number = ""
                else:
                    po_numbers = extract_po_from_areas(document_data)
                    po_number = ", ".join(po_numbers) if po_numbers else ""
                    
                    if po_number:
                        logger.info(f"Final PO Number(s): {po_number}")
                    else:
                        logger.warning("No PO number found")
        
        # Parse date safely
        parsed_date = ""
        if final_submission_date:
            date_str = str(final_submission_date).split(" ")[0]  # Remove time part
            date_str = date_str.replace("/", "-")  # Normalize slashes to dashes
            
            for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
                try:
                    parsed_date = datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue

        mapping = {
            "INTERNAL ID -1": data.get("uuid", ""),
            "INTERNAL ID -2": data.get("internalId", ""),
            "DATE": (
                parsed_date
                if parsed_date
                else (
                    datetime.strptime(
                        data.get("dateTimeReceived", "").split("T")[0],
                        "%Y-%m-%d",
                    ).strftime("%Y-%m-%d")
                    if data.get("dateTimeReceived")
                    else ""
                )
            ),
            "TYPE": mapped_type,
            "version": data.get("typeVersionName", ""),
            "TOTAL VALUE EGP": data.get("total", ""),
            "FROM": final_from,
            "REGESTRAION NUMBER": data.get("issuerId", ""),
            "STATUS": data.get("status", ""),
            "REGESTRAION": data.get("receiverId", ""),
            "PO number": po_number,
            # "filename": file_path.name
        }
        logger.info(f"📄 Extracted UUID: {data.get('uuid', 'MISSING')} from file: {file_path.name}")

        return mapping
        
    except json.JSONDecodeError:
        logger.error(f"JSON decode error in: {file_path.name}")
        return {
            "INTERNAL ID -1": "",
            "INTERNAL ID -2": "",
            "DATE": "",
            "TYPE": "Error",
            "version": "",
            "TOTAL VALUE EGP": "",
            "FROM": "",
            "REGESTRAION NUMBER": "",
            "STATUS": "JSON Error",
            "REGESTRAION": "",
            "PO number": "",
            # "filename": file_path.name
        }
    except Exception as e:
        logger.error(f"Error processing {file_path.name}: {e}")
        return {
            "INTERNAL ID -1": "",
            "INTERNAL ID -2": "",
            "DATE": "",
            "TYPE": "Error", 
            "version": "",
            "TOTAL VALUE EGP": "",
            "FROM": "",
            "REGESTRAION NUMBER": "",
            "STATUS": f"Error: {str(e)}",
            "REGESTRAION": "",
            "PO number": "",
            # "filename": file_path.name
        }

def save_to_excel(results, output_file):
    """Save results to Excel file"""
    try:
        from openpyxl.styles import Font
        
        df = pd.DataFrame(results)
        
        column_order = [
            # "filename",
              "INTERNAL ID -1", "INTERNAL ID -2", "DATE", "TYPE", 
            "version", "TOTAL VALUE EGP", "FROM", "REGESTRAION NUMBER", 
            "STATUS", "REGESTRAION", "PO number"
        ]
        
        df = df[column_order]
        
        # ✅ فحص التكرار في UUID (حماية إضافية)
        initial_count = len(df)
        duplicates = df[df.duplicated(subset=['INTERNAL ID -1'], keep=False)]
        
        if not duplicates.empty:
            logger.warning(f"⚠️ Found {len(duplicates)} rows with duplicate UUIDs:")
            for idx, row in duplicates.iterrows():
                logger.warning(f"   UUID: {row['INTERNAL ID -1']}, File: {row['filename']}")
            
            # احذف التكرارات وخلي أول واحدة بس
            df = df.drop_duplicates(subset=['INTERNAL ID -1'], keep='first')
            logger.warning(f"⚠️ Removed {initial_count - len(df)} duplicate rows. Final count: {len(df)}")
        
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # # ✨ إضافة الهايبرلينكس للفواتير
        # from openpyxl import load_workbook
        # wb = load_workbook(output_file)
        # ws = wb.active
        
        # base_url = "https://invoicing.eta.gov.eg/documents/"
        
        # # Loop through rows (starting from row 2 to skip header)
        # for row_num in range(2, ws.max_row + 1):
        #     invoice_id_cell = ws.cell(row=row_num, column=2)  # INTERNAL ID -1 column
        #     filename_cell = ws.cell(row=row_num, column=1)    # filename column
            
        #     invoice_id = invoice_id_cell.value
            
        #     if invoice_id:
        #         # Create the full URL
        #         invoice_url = f"{base_url}{invoice_id}"
                
        #         # Add hyperlink to filename cell
        #         filename_cell.hyperlink = invoice_url
        #         filename_cell.font = Font(color="0563C1", underline="single")
        #         filename_cell.style = "Hyperlink"
        
        # wb.save(output_file)
        # logger.info(f"✅ Added hyperlinks to {ws.max_row - 1} invoice filenames")
        logger.info(f"Results saved to: {output_file}")
        logger.info(f"Total records: {len(df)}")
        logger.info(f"Records with PO numbers: {len(df[df['PO number'] != ''])}")
        logger.info(f"Records with issuer names from Excel: {len(df[df['FROM'] != ''])}")
        logger.info(f"Records with submission dates: {len(df[df['SUBMISSION DATE'] != ''])}")
        return True
    except Exception as e:
        logger.error(f"Error saving Excel: {e}")
        return False

def copy_pdfs_to_output(date_str, outputs_date_dir):
    """Copy all PDFs from invoices_pdf/<date> to outputs/<date>/PDF/ and delete source folder"""
    pdf_source_dir = Path("invoices_pdf") / date_str
    pdf_output_dir = outputs_date_dir / "PDF"
    
    if not pdf_source_dir.exists():
        logger.warning(f"PDF source directory does not exist: {pdf_source_dir}")
        return True
    
    try:
        pdf_output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Created PDF output directory: {pdf_output_dir}")
        
        total_pdfs_copied = 0
        
        for supplier_folder in pdf_source_dir.iterdir():
            if supplier_folder.is_dir():
                supplier_name = supplier_folder.name

                # ✅ توحيد أسماء الموردين
                if "شركه ثري ام بي" in supplier_name:
                    supplier_name = "3MP"
                elif "مكتب علمي ام ام فارما" in supplier_name:
                    supplier_name = "MMP"

                supplier_pdf_output_dir = pdf_output_dir / supplier_name
                supplier_pdf_output_dir.mkdir(parents=True, exist_ok=True)
                
                pdf_files = list(supplier_folder.glob('*.pdf'))
                if pdf_files:
                    logger.info(f"Copying {len(pdf_files)} PDFs for supplier: {supplier_name}")
                    
                    for pdf_file in pdf_files:
                        destination = supplier_pdf_output_dir / pdf_file.name
                        shutil.copy2(pdf_file, destination)
                        logger.debug(f"Copied PDF: {pdf_file.name} → {destination}")
                        total_pdfs_copied += 1
                    
                    logger.info(f"Successfully copied {len(pdf_files)} PDFs to: {supplier_pdf_output_dir}")
                else:
                    logger.warning(f"No PDF files found for supplier: {supplier_name}")
        
        if total_pdfs_copied > 0:
            logger.info(f"Total PDFs copied: {total_pdfs_copied}")
            logger.info(f"Deleting source PDF directory: {pdf_source_dir}")
            shutil.rmtree(pdf_source_dir)
            logger.info(f"Successfully deleted source PDF directory: {pdf_source_dir}")
        else:
            logger.warning("No PDFs were copied, keeping source directory")
        
        return True
        
    except Exception as e:
        logger.error(f"Error copying PDFs: {e}")
        return False

def process_taxpayer(taxpayer_folder, outputs_date_dir, issuer_data_dict):
    """Process all JSON files for a single taxpayer"""
    taxpayer_name = taxpayer_folder.name

    # ✅ Replace Arabic supplier names with English codes
    if "شركه ثري ام بي" in taxpayer_name:
        taxpayer_name = "3MP"
    elif "مكتب علمي ام ام فارما" in taxpayer_name:
        taxpayer_name = "MMP"

    logger.info(f"Processing taxpayer: {taxpayer_name}")
    
    taxpayer_excel_output_dir = outputs_date_dir / "Excel" / taxpayer_name
    taxpayer_excel_output_dir.mkdir(parents=True, exist_ok=True)
    
    json_files = list(taxpayer_folder.glob('*.json'))
    if not json_files:
        logger.warning(f"No JSON files found for {taxpayer_name}")
        return False
    
    results = []
    seen_uuids = set()  # ✅ تتبع الـ UUIDs المعالجة
    
    for i, json_file in enumerate(json_files, 1):
        logger.info(f"[{i}/{len(json_files)}] Processing {json_file.name}")
        result = process_json_file(json_file, issuer_data_dict)
        
        if result:
            uuid = result.get("INTERNAL ID -1", "")
            
            # ✅ تحقق من التكرار
            if uuid and uuid in seen_uuids:
                logger.warning(f"⚠️ DUPLICATE UUID SKIPPED: {uuid} in file {json_file.name}")
                continue
            
            # ✅ تحقق من وجود UUID
            if not uuid:
                logger.error(f"❌ MISSING UUID in file {json_file.name} - SKIPPING")
                continue
            
            # ✅ أضف UUID للقائمة المعالجة
            seen_uuids.add(uuid)
            results.append(result)
            logger.info(f"✅ Added invoice: {uuid}")
    
    if not results:
        logger.warning(f"No valid results to save for {taxpayer_name} (all files excluded)")
        return False
    
    output_file = taxpayer_excel_output_dir / "results.xlsx"
    if save_to_excel(results, output_file):
        logger.info(f"Successfully processed {taxpayer_name} - Excel saved to: {output_file}")
        return True
    else:
        logger.error(f"Failed to save Excel results for {taxpayer_name}")
        return False

def cleanup_old_logs():
    """Clean up old daily log files, keep only the main json_parser.log"""
    log_dir = Path("logs")
    if not log_dir.exists():
        return
    
    for log_file in log_dir.glob("json_parser_*.log"):
        try:
            log_file.unlink()
            print(f"Deleted old daily log: {log_file.name}")
        except Exception as e:
            print(f"Could not delete log file {log_file.name}: {e}")

def main():
    """Main execution function"""
    cleanup_old_logs()
    
    logger.info("Starting JSON parsing process")
    logger.info(f"Main log file: logs/json_parser.log (overwritten each run)")
    logger.info("=" * 80)
    logger.info("IMPROVED PO EXTRACTION FEATURES:")
    logger.info("✅ Skip postal codes: 1111, 12345, etc.")
    logger.info("✅ Skip vehicle numbers: chassis, plate numbers")
    logger.info("✅ Skip delivery/product descriptions without PO keywords")
    logger.info("✅ Skip manufacturing years in context")
    logger.info("✅ Extract PO before registration number")
    logger.info("✅ Only extract from invoice lines with PO keywords")
    logger.info("🟩 Case 1: Use JSON issuer.name when Excel = 'غير محدد' or empty")
    logger.info("🟨 Case 2: Exclude suppliers (مكتب علمي ام ام فارما, شركه ثري ام بي) from output")
    logger.info("🟦 Case 3: Barakat Group - Extract PO from JSON fields directly")
    logger.info("📅 NEW: Read Submission Date from Excel scraping file")
    logger.info("=" * 80)
    
    logger.info("Loading issuer data (names + submission dates) from Excel file...")
    issuer_data_dict = load_issuer_data_from_excel()
    
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--date', type=str)
    args = parser.parse_args()

    if args.date:
        yesterday = args.date
    else:
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%d-%m-%Y")
    
    base_path = Path("invoices_json") / yesterday
    outputs_path = Path("outputs")
    outputs_path.mkdir(exist_ok=True)
    
    outputs_date_dir = outputs_path / yesterday
    outputs_date_dir.mkdir(exist_ok=True)
    
    if not base_path.exists():
        logger.error(f"Yesterday's folder does not exist: {base_path}")
        sys.exit(1)
    
    logger.info(f"Looking for taxpayer folders in: {base_path}")
    logger.info(f"Output structure will be: {outputs_date_dir}/Excel/<SupplierName>/results.xlsx")
    logger.info(f"PDF structure will be: {outputs_date_dir}/PDF/<SupplierName>/*.pdf")
    
    successful_taxpayers = 0
    failed_taxpayers = 0
    
    for taxpayer_folder in base_path.iterdir():
        if taxpayer_folder.is_dir():
            try:
                if process_taxpayer(taxpayer_folder, outputs_date_dir, issuer_data_dict):
                    successful_taxpayers += 1
                else:
                    failed_taxpayers += 1
            except Exception as e:
                logger.error(f"Error processing taxpayer {taxpayer_folder.name}: {e}")
                failed_taxpayers += 1
    
    logger.info("Starting PDF copy process...")
    pdf_copy_success = copy_pdfs_to_output(yesterday, outputs_date_dir)
    
    logger.info("=" * 80)
    logger.info("JSON parsing process completed")
    logger.info(f"Successful taxpayers: {successful_taxpayers}")
    logger.info(f"Failed taxpayers: {failed_taxpayers}")
    logger.info(f"PDF copy successful: {pdf_copy_success}")
    logger.info(f"Output directory structure:")
    logger.info(f"  - Excel files: {outputs_date_dir}/Excel/<SupplierName>/results.xlsx")
    logger.info(f"  - PDF files: {outputs_date_dir}/PDF/<SupplierName>/*.pdf")
    logger.info(f"Issuer data loaded from Excel: {len(issuer_data_dict)} records")
    logger.info("=" * 80)
    
    if successful_taxpayers > 0:
        logger.info(f"Results saved in: {outputs_date_dir}")
        print(f"\n✅ JSON parsing completed successfully!")
        print(f"📂 Output folder: {outputs_date_dir.absolute()}")
        print(f"📊 Excel files: {outputs_date_dir}/Excel/<SupplierName>/results.xlsx")
        print(f"📄 PDF files: {outputs_date_dir}/PDF/<SupplierName>/*.pdf")
        print(f"✅ Processed {successful_taxpayers} taxpayers")
        print(f"📝 Issuer data loaded from Excel: {len(issuer_data_dict)} records")
        print(f"\n🆕 IMPROVED FEATURES:")
        print(f"✅ Skip postal codes (1111, 12345, etc.)")
        print(f"✅ Skip vehicle numbers (chassis, plates)")
        print(f"✅ Skip delivery descriptions without PO keywords")
        print(f"✅ Skip manufacturing years")
        print(f"✅ Extract PO before registration number")
        print(f"🟩 Excel 'غير محدد' → Use JSON issuer.name")
        print(f"🟨 Excluded suppliers: مكتب علمي ام ام فارما, شركه ثري ام بي")
        print(f"🟦 Barakat Group → PO from JSON fields")
        print(f"📅 NEW: Submission Date from Excel scraping file")
        if failed_taxpayers > 0:
            print(f"❌ Failed {failed_taxpayers} taxpayers")
        if pdf_copy_success:
            print(f"📁 PDFs successfully copied and source folder deleted")
        else:
            print(f"⚠️ Warning: PDF copy process had issues")
    else:
        logger.error("No taxpayers were processed successfully")
        sys.exit(1)

if __name__ == "__main__":
    main()