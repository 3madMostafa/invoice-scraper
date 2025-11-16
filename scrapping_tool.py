from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time, os, logging, shutil
from datetime import datetime, timedelta
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configuration
EMAIL = "kamal_hanna@globalnapi.com"
PASSWORD = "Pola123$"

# Taxpayers mapping (Name → Value)
TAXPAYERS = {
    "مكتب علمي ام ام فارما": "54037",
    "شركه ثري ام بي": "54041"
}

# Fixed target date
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('--date', type=str)
args = parser.parse_args()

if args.date:
    TARGET_DATE = datetime.strptime(args.date, "%d-%m-%Y").date()
    TARGET_DATE_STR = args.date
else:
    TARGET_DATE = (datetime.now() - timedelta(days=1)).date()
    TARGET_DATE_STR = TARGET_DATE.strftime("%d-%m-%Y")

# Setup centralized logging
def setup_logging():
    log_dir = os.path.join(os.getcwd(), "logs")
    os.makedirs(log_dir, exist_ok=True)
    
    log_filename = 'scraping.log'
    log_filepath = os.path.join(log_dir, log_filename)
    
    should_reset = True
    if os.path.exists(log_filepath):
        try:
            file_creation_time = datetime.fromtimestamp(os.path.getctime(log_filepath))
            today = datetime.now().date()
            file_date = file_creation_time.date()
            
            if file_date == today:
                should_reset = False
                print(f"Appending to today's log: {log_filepath}")
            else:
                print(f"Log is from {file_date}, creating fresh log for {today}")
        except Exception as e:
            print(f"Error checking log file date: {e}")
    
    logger = logging.getLogger('scraper')
    logger.setLevel(logging.INFO)
    
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    file_handler = logging.FileHandler(
        log_filepath, 
        encoding='utf-8',
        mode='w' if should_reset else 'a'
    )
    file_handler.setLevel(logging.INFO)
    
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    current_time = datetime.now().strftime("%H:%M:%S")
    if should_reset:
        logger.info(f"📅 NEW DAY LOG STARTED - {datetime.now().strftime('%Y-%m-%d')}")
        logger.info(f"=" * 60)
    logger.info(f"🚀 NEW SCRAPING RUN STARTED AT {current_time}")
    logger.info(f"=" * 60)
    
    return logger

def setup_excel_file():
    """Setup Excel file for invoices data"""
    try:
        excel_dir = os.path.join(os.getcwd(), "logs")
        os.makedirs(excel_dir, exist_ok=True)
        
        excel_filename = f'invoices_data_{TARGET_DATE_STR}.xlsx'
        excel_filepath = os.path.join(excel_dir, excel_filename)
        
        if os.path.exists(excel_filepath):
            wb = load_workbook(excel_filepath)
            ws = wb.active
            logger.info(f"Loading existing Excel file: {excel_filepath}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices Data"
            
            headers = ['Invoice ID', 'Issuer Name', 'Submission Date', 'Status', 'Taxpayer', 'Date Processed']
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 20
            
            wb.save(excel_filepath)
            logger.info(f"Created new Excel file: {excel_filepath}")
        
        return excel_filepath, wb, ws
        
    except Exception as e:
        logger.error(f"Error setting up Excel file: {e}")
        return None, None, None

def add_invoice_to_excel(invoice_id, issuer_name, submission_date, status, taxpayer_name):
    """Add invoice data to Excel file - SAVES IMMEDIATELY"""
    try:
        invoice_data = {
            'invoice_id': invoice_id,
            'issuer_name': issuer_name,
            'submission_date': submission_date,
            'status': status,
            'taxpayer': taxpayer_name,
            'date_processed': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        scraping_summary['invoices_data'].append(invoice_data)
        
        # SAVE TO EXCEL IMMEDIATELY - كل invoice لوحده
        excel_filepath, wb, ws = setup_excel_file()
        if excel_filepath:
            row_data = [
                invoice_data['invoice_id'],
                invoice_data['issuer_name'],
                invoice_data['submission_date'],
                invoice_data['status'],
                invoice_data['taxpayer'],
                invoice_data['date_processed']
            ]
            ws.append(row_data)
            
            wb.save(excel_filepath)
            logger.info(f"✓ Saved to Excel: {invoice_id}")
        
    except Exception as e:
        logger.error(f"Error adding invoice to Excel data: {e}")

def save_invoices_to_excel():
    """Save all collected invoice data to Excel file"""
    try:
        if not scraping_summary['invoices_data']:
            logger.info("No invoice data to save to Excel")
            return
        
        excel_filepath, wb, ws = setup_excel_file()
        if not excel_filepath:
            logger.error("Failed to setup Excel file")
            return
        
        for invoice_data in scraping_summary['invoices_data']:
            row_data = [
                invoice_data['invoice_id'],
                invoice_data['issuer_name'],
                invoice_data['submission_date'],
                invoice_data['status'],
                invoice_data['taxpayer'],
                invoice_data['date_processed']
            ]
            ws.append(row_data)
        
        for row_num in range(ws.max_row - len(scraping_summary['invoices_data']) + 1, ws.max_row + 1):
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if col_num == 4:  # Status column is now column 4
                    if cell.value == "Downloaded":
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif cell.value == "Partial Download":
                        cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                    elif cell.value == "Failed":
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    elif cell.value == "Cancelled":
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        wb.save(excel_filepath)
        logger.info(f"Successfully saved {len(scraping_summary['invoices_data'])} invoices to Excel: {excel_filepath}")
        
    except Exception as e:
        logger.error(f"Error saving invoices to Excel: {e}")

logger = setup_logging()

# Global variables for summary
scraping_summary = {
    'start_time': None,
    'end_time': None,
    'taxpayers_processed': [],
    'total_invoices_downloaded': 0,
    'cancelled_invoices': 0,
    'errors': [],
    'successful_downloads': {'json': 0, 'pdf': 0},
    'invoices_data': []
}

# Setup directories
base_dir = os.getcwd()
json_root = os.path.join(base_dir, "invoices_json")
pdf_root = os.path.join(base_dir, "invoices_pdf")

json_date_dir = os.path.join(json_root, TARGET_DATE_STR)
pdf_date_dir = os.path.join(pdf_root, TARGET_DATE_STR)

os.makedirs(json_date_dir, exist_ok=True)
os.makedirs(pdf_date_dir, exist_ok=True)

current_json_dir = ""
current_pdf_dir = ""

# Optimized Chrome options
options = webdriver.ChromeOptions()

# Create a default download directory
default_download_dir = os.path.join(base_dir, "temp_downloads")
os.makedirs(default_download_dir, exist_ok=True)

options.add_experimental_option("prefs", {
    "download.default_directory": default_download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": False,
    "safebrowsing.disable_download_protection": True,
    "profile.default_content_settings.popups": 0,
    "profile.content_settings.exceptions.automatic_downloads.*.setting": 1
})
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")
options.add_argument("--disable-extensions")
options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-blink-features=AutomationControlled")

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 20)

# Enable download behavior
driver.execute_cdp_cmd("Page.setDownloadBehavior", {
    "behavior": "allow",
    "downloadPath": default_download_dir
})

def create_taxpayer_directories(taxpayer_name):
    """Create directories for specific taxpayer"""
    global current_json_dir, current_pdf_dir
    
    current_json_dir = os.path.join(json_date_dir, taxpayer_name)
    current_pdf_dir = os.path.join(pdf_date_dir, taxpayer_name)
    
    os.makedirs(current_json_dir, exist_ok=True)
    os.makedirs(current_pdf_dir, exist_ok=True)
    
    logger.info(f"Created directories for {taxpayer_name}")

def wait_for_page_load(timeout=15):
    """Wait for page to load - optimized"""
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        
        try:
            WebDriverWait(driver, 3).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ms-Overlay, .loading, .spinner"))
            )
        except TimeoutException:
            pass
        
        time.sleep(0.8)
        return True
    except:
        time.sleep(1)
        return False

def wait_overlay_to_disappear(timeout=8):
    """Wait for overlay to disappear"""
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ms-Overlay"))
        )
        time.sleep(0.3)
    except:
        time.sleep(0.5)

def safe_click(element_locator, timeout=12, description="element"):
    """Safely click an element"""
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(element_locator)
        )
        
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        time.sleep(0.2)
        
        try:
            element.click()
        except:
            driver.execute_script("arguments[0].click();", element)
        
        time.sleep(0.3)
        return True
    except:
        return False

def safe_send_keys(element_locator, text, timeout=12, description="element"):
    """Safely send keys to an element"""
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(element_locator)
        )
        element.clear()
        time.sleep(0.2)
        element.send_keys(text)
        time.sleep(0.2)
        return True
    except:
        return False

def wait_for_rows_to_load(timeout=15):
    """Wait for table rows to load"""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[role='row']"))
        )
        
        WebDriverWait(driver, timeout).until(
            lambda d: len([
                cell for cell in d.find_elements(By.CSS_SELECTOR, "div[data-automation-key='dateTimeReceived']")
                if cell.text.strip()
            ]) > 0
        )
        
        time.sleep(0.8)
        return True
    except:
        time.sleep(1)
        return False

def click_next_page():
    """Click the next page button"""
    try:
        wait_overlay_to_disappear()
        
        next_btn_xpath = "//i[contains(@data-icon-name,'ChevronRight')]/ancestor::button"
        
        next_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, next_btn_xpath))
        )
        
        if next_btn.get_attribute("disabled") or "is-disabled" in next_btn.get_attribute("class"):
            return False
        
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
        time.sleep(0.3)
        
        try:
            next_btn.click()
        except:
            driver.execute_script("arguments[0].click();", next_btn)
        
        time.sleep(1.2)
        wait_for_page_load()
        wait_for_rows_to_load()
        
        return True
    except:
        return False

def check_if_cancelled():
    """Check if invoice is cancelled or rejected - returns status"""
    try:
        time.sleep(0.5)
        
        # Check for cancelled or rejected indicators
        cancelled_selectors = [
            # cancelled
            "div.horizontal.cancelled",
            "div.cancelled",
            "div[class*='cancelled']",
            "//span[contains(text(), 'Cancelled')]",
            "//span[contains(text(), 'ملغاة')]",

            # rejected
            "div.horizontal.valid-rejected",
            "div[class*='valid-rejected']",
            "//span[contains(text(), 'Rejected')]",
            "//span[contains(text(), 'مرفوضة')]"
        ]
        
        for selector in cancelled_selectors:
            try:
                if selector.startswith("//"):
                    elements = driver.find_elements(By.XPATH, selector)
                else:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                
                for elem in elements:
                    if elem.is_displayed():
                        elem_text = elem.text.strip().lower()
                        if 'cancel' in elem_text or 'ملغ' in elem_text:
                            logger.info("⚠️ Invoice is CANCELLED")
                            return "Cancelled"
                        if 'reject' in elem_text or 'مرفوض' in elem_text:
                            logger.info("⚠️ Invoice is REJECTED")
                            return "Rejected"
            except:
                continue
        
        return None  # Valid invoice
        
    except:
        return None

def get_issuer_name(max_attempts=10):
    """Extract issuer name - CRITICAL with 6 attempts"""
    
    for attempt in range(max_attempts):
        try:
            time.sleep(0.8)
            
            # Wait for readonly inputs
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.ms-TextField.eta-cTextFieldReadOnly input[readonly]"))
                )
                time.sleep(0.3)
            except TimeoutException:
                if attempt < max_attempts - 1:
                    logger.warning(f"Inputs timeout (attempt {attempt + 1}), refreshing...")
                    driver.refresh()
                    wait_for_page_load()
                    time.sleep(1)
                    continue
            
            # Get inputs with multiple selectors
            selectors = [
                "input[readonly]",
                "div.ms-TextField.eta-cTextFieldReadOnly input[readonly]",
                "input.ms-TextField-field[readonly]",
                "input[readonly][type='text']"
            ]
            
            issuer_inputs = []
            for selector in selectors:
                try:
                    issuer_inputs.extend(driver.find_elements(By.CSS_SELECTOR, selector))
                except:
                    continue
            
            if not issuer_inputs:
                if attempt < max_attempts - 1:
                    logger.warning(f"No inputs (attempt {attempt + 1}), refreshing...")
                    driver.refresh()
                    wait_for_page_load()
                    time.sleep(1)
                    continue
                return "غير محدد"
            
            # Extract valid names
            potential_names = []
            
            for elem in issuer_inputs:
                try:
                    if not elem.is_displayed():
                        continue
                    
                    value = elem.get_attribute("value")
                    if not value or not value.strip():
                        continue
                    
                    value = value.strip()
                    
                    # Filter invalid values
                    if (len(value) < 3 or 
                        value.isdigit() or 
                        "@" in value or
                        "http" in value.lower() or
                        value.startswith("EGP") or
                        value.startswith("$") or
                        value.replace(".", "").replace(",", "").isdigit()):
                        continue
                    
                    potential_names.append(value)
                    
                except:
                    continue
            
            if potential_names:
                issuer_name = potential_names[0]
                logger.info(f"✓ Issuer: {issuer_name}")
                return issuer_name
            
            # Try label-based extraction
            try:
                labels = driver.find_elements(By.CSS_SELECTOR, "label.ms-Label")
                for label in labels:
                    label_text = label.text.strip()
                    if any(kw in label_text for kw in ["اسم", "Name", "Issuer"]):
                        try:
                            parent = label.find_element(By.XPATH, "./parent::*")
                            inputs = parent.find_elements(By.CSS_SELECTOR, "input")
                            for inp in inputs:
                                val = inp.get_attribute("value")
                                if val and len(val.strip()) > 3:
                                    logger.info(f"✓ Issuer (label): {val.strip()}")
                                    return val.strip()
                        except:
                            continue
            except:
                pass
            
            if attempt < max_attempts - 1:
                logger.warning(f"No valid issuer (attempt {attempt + 1}), refreshing...")
                driver.refresh()
                wait_for_page_load()
                time.sleep(1)
                continue
            
        except Exception as e:
            if attempt < max_attempts - 1:
                logger.warning(f"Error attempt {attempt + 1}, refreshing...")
                driver.refresh()
                wait_for_page_load()
                time.sleep(1)
                continue
    
    logger.warning("⚠ Could not extract issuer")
    return "غير محدد"

def get_issuer_name(max_attempts=6):
    """Extract issuer name and submission date - CRITICAL with 6 attempts"""
    
    for attempt in range(max_attempts):
        try:
            time.sleep(0.8)
            
            # Wait for readonly inputs
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.ms-TextField.eta-cTextFieldReadOnly input[readonly]"))
                )
                time.sleep(0.3)
            except TimeoutException:
                if attempt < max_attempts - 1:
                    logger.warning(f"Inputs timeout (attempt {attempt + 1}), refreshing...")
                    driver.refresh()
                    wait_for_page_load()
                    time.sleep(1)
                    continue
            
            # Get inputs with multiple selectors
            selectors = [
                "input[readonly]",
                "div.ms-TextField.eta-cTextFieldReadOnly input[readonly]",
                "input.ms-TextField-field[readonly]",
                "input[readonly][type='text']"
            ]
            
            issuer_inputs = []
            for selector in selectors:
                try:
                    issuer_inputs.extend(driver.find_elements(By.CSS_SELECTOR, selector))
                except:
                    continue
            
            if not issuer_inputs:
                if attempt < max_attempts - 1:
                    logger.warning(f"No inputs (attempt {attempt + 1}), refreshing...")
                    driver.refresh()
                    wait_for_page_load()
                    time.sleep(1)
                    continue
                return "غير محدد", "غير محدد"
            
            # Extract valid names
            potential_names = []
            
            for elem in issuer_inputs:
                try:
                    if not elem.is_displayed():
                        continue
                    
                    value = elem.get_attribute("value")
                    if not value or not value.strip():
                        continue
                    
                    value = value.strip()
                    
                    # Filter invalid values
                    if (len(value) < 3 or 
                        value.isdigit() or 
                        "@" in value or
                        "http" in value.lower() or
                        value.startswith("EGP") or
                        value.startswith("$") or
                        value.replace(".", "").replace(",", "").isdigit()):
                        continue
                    
                    potential_names.append(value)
                    
                except:
                    continue
            
            issuer_name = "غير محدد"
            if potential_names:
                issuer_name = potential_names[0]
            else:
                # Try label-based extraction
                try:
                    labels = driver.find_elements(By.CSS_SELECTOR, "label.ms-Label")
                    for label in labels:
                        label_text = label.text.strip()
                        if any(kw in label_text for kw in ["اسم", "Name", "Issuer"]):
                            try:
                                parent = label.find_element(By.XPATH, "./parent::*")
                                inputs = parent.find_elements(By.CSS_SELECTOR, "input")
                                for inp in inputs:
                                    val = inp.get_attribute("value")
                                    if val and len(val.strip()) > 3:
                                        issuer_name = val.strip()
                                        break
                            except:
                                continue
                except:
                    pass
            
            # Extract Submission Date
            submission_date = "غير محدد"
            try:
                # Try to find submission date in the specific div
                submission_divs = driver.find_elements(By.XPATH, "//div[@class='flex']//div[@class='OnDesktopView']")
                for div in submission_divs:
                    div_text = div.text.strip()
                    if "Submission Date:" in div_text:
                        # Extract the date part
                        date_parts = div_text.split("Submission Date:")
                        if len(date_parts) > 1:
                            date_str = date_parts[1].strip()
                            # Take only the first part before (UTC)
                            if "(" in date_str:
                                date_str = date_str.split("(")[0].strip()
                            submission_date = date_str
                            break
                
                # Alternative method if first method fails
                if submission_date == "غير محدد":
                    span_elements = driver.find_elements(By.XPATH, "//span[contains(text(), 'Submission Date:')]/following-sibling::span")
                    if span_elements:
                        date_text = span_elements[0].text.strip()
                        if "(" in date_text:
                            date_text = date_text.split("(")[0].strip()
                        submission_date = date_text
                        
            except Exception as e:
                logger.warning(f"Could not extract submission date: {e}")
            
            if issuer_name != "غير محدد" or submission_date != "غير محدد":
                logger.info(f"✓ Issuer: {issuer_name} | Submission: {submission_date}")
                return issuer_name, submission_date
            
            if attempt < max_attempts - 1:
                logger.warning(f"No valid issuer/date (attempt {attempt + 1}), refreshing...")
                driver.refresh()
                wait_for_page_load()
                time.sleep(1)
                continue
            
        except Exception as e:
            if attempt < max_attempts - 1:
                logger.warning(f"Error attempt {attempt + 1}, refreshing...")
                driver.refresh()
                wait_for_page_load()
                time.sleep(1)
                continue
    
    logger.warning("⚠ Could not extract issuer/submission date")
    return "غير محدد", "غير محدد"

def move_downloaded_file(source_dir, target_dir, file_extension, invoice_id):
    """Move downloaded file"""
    try:
        time.sleep(0.8)
        
        if not os.path.exists(source_dir):
            return False
        
        files = [f for f in os.listdir(source_dir) if f.endswith(file_extension)]
        
        if not files:
            return False
        
        files_with_time = [(f, os.path.getmtime(os.path.join(source_dir, f))) for f in files]
        files_with_time.sort(key=lambda x: x[1], reverse=True)
        latest_file = files_with_time[0][0]
        
        source_path = os.path.join(source_dir, latest_file)
        target_filename = f"{invoice_id}{file_extension}"
        target_path = os.path.join(target_dir, target_filename)
        
        if os.path.exists(source_path):
            if os.path.exists(target_path):
                os.remove(target_path)
            
            shutil.move(source_path, target_path)
            return True
        return False
            
    except:
        return False

def wait_for_download_complete(download_dir, expected_extension, timeout=20):
    """Wait for download to complete"""
    start_time = time.time()
    last_check_time = start_time
    last_file_count = 0
    
    try:
        while time.time() - start_time < timeout:
            try:
                if not os.path.exists(download_dir):
                    time.sleep(0.2)
                    continue
                    
                files = os.listdir(download_dir)
                
                temp_files = [f for f in files if f.endswith('.crdownload') or f.endswith('.tmp')]
                if temp_files:
                    last_check_time = time.time()
                    time.sleep(0.2)
                    continue
                
                target_files = [f for f in files if f.lower().endswith(expected_extension.lower())]
                
                if len(target_files) > last_file_count:
                    time.sleep(1)
                    return True
                
                last_file_count = len(target_files)
                
                if time.time() - last_check_time > 7 and not temp_files:
                    return False
                
                time.sleep(0.2)
                
            except:
                time.sleep(0.2)
        
        return False
        
    except:
        return False

def download_file(file_type, invoice_id, max_retries=5):
    """
    Download file with improved retry logic - CRITICAL FUNCTION
    Only refresh after actual failure, not at start of each attempt
    """
    temp_download_dir = os.path.join(base_dir, "temp_downloads")
    final_dir = current_pdf_dir if file_type == 'PDF' else current_json_dir
    expected_extension = '.pdf' if file_type == 'PDF' else '.json'
    
    # Clear temp directory once at the start
    try:
        for f in os.listdir(temp_download_dir):
            if f.endswith(expected_extension):
                os.remove(os.path.join(temp_download_dir, f))
    except:
        pass
    
    needs_refresh = False  # Track if we need to refresh
    
    for attempt in range(max_retries):
        try:
            logger.info(f"Downloading {file_type} - attempt {attempt + 1}/{max_retries}")
            
            # Only refresh if previous attempt actually failed
            if needs_refresh:
                logger.info(f"🔄 Refreshing page after failure (attempt {attempt + 1})")
                driver.refresh()
                wait_for_page_load()
                time.sleep(1.5)
                
                # Check if cancelled after refresh
                if check_if_cancelled():
                    logger.info("Invoice cancelled after refresh")
                    return False
                
                needs_refresh = False  # Reset flag
            
            # Small wait before starting
            wait_overlay_to_disappear()
            time.sleep(0.5)
            
            # STEP 1: Find and click Download button
            download_btn_xpath = "//button[contains(., 'Download as')]"
            
            try:
                download_btn = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, download_btn_xpath))
                )
                
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", download_btn)
                time.sleep(0.3)
                
                download_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, download_btn_xpath))
                )
                
                try:
                    download_btn.click()
                except:
                    driver.execute_script("arguments[0].click();", download_btn)
                
                logger.info("✓ Clicked Download button")
                time.sleep(1.2)
                
            except Exception as e:
                logger.warning(f"Failed to click Download button: {str(e)[:80]}")
                if attempt < max_retries - 1:
                    needs_refresh = True  # Mark that we need refresh
                    continue
                else:
                    raise Exception("Download button not found")
            
            # STEP 2: Wait for dropdown to appear
            dropdown_appeared = False
            for wait_try in range(5):
                try:
                    dropdown_options = driver.find_elements(By.XPATH, f"//span[contains(text(), '{file_type}')]")
                    visible_options = [opt for opt in dropdown_options if opt.is_displayed()]
                    if visible_options:
                        dropdown_appeared = True
                        logger.info("✓ Dropdown appeared")
                        break
                except:
                    pass
                time.sleep(0.4)
            
            if not dropdown_appeared:
                logger.warning("Dropdown did not appear")
                if attempt < max_retries - 1:
                    needs_refresh = True  # Mark that we need refresh
                    continue
                else:
                    raise Exception("Dropdown not visible")
            
            # STEP 3: Click file type option
            xpath_patterns = [
                f"//span[text()='{file_type}']",
                f"//span[contains(text(), '{file_type}')]",
                f"//button[contains(., '{file_type}')]"
            ]
            
            file_option = None
            for pattern in xpath_patterns:
                try:
                    elements = driver.find_elements(By.XPATH, pattern)
                    visible = [e for e in elements if e.is_displayed() and e.is_enabled()]
                    if visible:
                        file_option = visible[0]
                        break
                except:
                    continue
            
            if not file_option:
                logger.warning(f"{file_type} option not found")
                if attempt < max_retries - 1:
                    needs_refresh = True  # Mark that we need refresh
                    continue
                else:
                    raise NoSuchElementException(f"{file_type} option not found")
            
            # Click the option
            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", file_option)
                time.sleep(0.2)
                
                try:
                    file_option.click()
                except:
                    driver.execute_script("arguments[0].click();", file_option)
                
                logger.info(f"✓ Clicked {file_type} option")
                time.sleep(0.5)
                
            except Exception as e:
                logger.warning(f"Failed to click {file_type} option: {str(e)[:80]}")
                if attempt < max_retries - 1:
                    needs_refresh = True  # Mark that we need refresh
                    continue
                else:
                    raise Exception(f"Failed to click {file_type} option")
            
            # STEP 4: Wait for download to complete
            if wait_for_download_complete(temp_download_dir, expected_extension, timeout=20):
                # STEP 5: Move file to final location
                if move_downloaded_file(temp_download_dir, final_dir, expected_extension, invoice_id):
                    logger.info(f"✓✓ {file_type} downloaded successfully")
                    scraping_summary['successful_downloads'][file_type.lower()] += 1
                    return True
                else:
                    logger.warning("Download succeeded but move failed")
                    if attempt < max_retries - 1:
                        needs_refresh = True  # Mark that we need refresh
                        continue
                    else:
                        raise Exception("File move failed")
            else:
                logger.warning("Download verification failed")
                if attempt < max_retries - 1:
                    needs_refresh = True  # Mark that we need refresh
                    continue
                else:
                    raise Exception("Download not detected")
            
        except Exception as e:
            logger.warning(f"Attempt {attempt + 1} error: {str(e)[:100]}")
            if attempt < max_retries - 1:
                needs_refresh = True  # Mark that we need refresh for next iteration
                time.sleep(0.5)
                continue
            else:
                error_msg = f"All {max_retries} attempts failed for {file_type} of {invoice_id}"
                logger.error(error_msg)
                scraping_summary['errors'].append(error_msg)
                return False
    
    return False

def download_both_files(invoice_id, taxpayer_name):
    """Download both JSON and PDF - including cancelled/rejected invoices"""
    # Check what files already exist
    json_file_path = os.path.join(current_json_dir, f"{invoice_id}.json")
    pdf_file_path = os.path.join(current_pdf_dir, f"{invoice_id}.pdf")
    
    json_exists = os.path.exists(json_file_path)
    pdf_exists = os.path.exists(pdf_file_path)
    
    # Check if we need to update issuer info in Excel
    need_issuer_update = False
    if json_exists and pdf_exists:
        try:
            excel_filepath, wb, ws = setup_excel_file()
            if excel_filepath:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == invoice_id:  # Column A = Invoice ID
                        issuer_name = row[1]  # Column B = Issuer Name
                        if issuer_name == "غير محدد":
                            need_issuer_update = True
                            logger.info(f"📝 Files exist but issuer='غير محدد' for {invoice_id}, will retrieve issuer info")
                        break
        except Exception as e:
            logger.warning(f"Could not check Excel for issuer: {e}")
        
        # If both files exist AND issuer is valid, skip completely
        if not need_issuer_update:
            logger.info(f"✓ Both files already exist for {invoice_id}, skipping download")
            return True
        else:
            # Files exist but we need issuer info only
            logger.info(f"🔍 Getting issuer info for {invoice_id} (files already downloaded)")
            
            try:
                invoice_status = check_if_cancelled()
                issuer_name, submission_date = get_issuer_name(max_attempts=6)
                
                # Update Excel with the new issuer info
                if issuer_name != "غير محدد":
                    # Delete old row and add new one
                    if excel_filepath:
                        rows_to_delete = []
                        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                            if row[0].value == invoice_id:
                                rows_to_delete.append(idx)
                        
                        for row_idx in sorted(rows_to_delete, reverse=True):
                            ws.delete_rows(row_idx)
                        
                        wb.save(excel_filepath)
                    
                    status_label = invoice_status if invoice_status else "Downloaded"
                    add_invoice_to_excel(invoice_id, issuer_name, submission_date, status_label, taxpayer_name)
                    logger.info(f"✓ Updated issuer info for {invoice_id}: {issuer_name}")
                    return True
                else:
                    logger.warning(f"⚠️ Still couldn't get issuer for {invoice_id}")
                    return True
                    
            except Exception as e:
                logger.error(f"Error updating issuer info: {e}")
                return True
    
    if json_exists:
        logger.info(f"📄 JSON exists for {invoice_id}, will only download PDF")
    if pdf_exists:
        logger.info(f"📄 PDF exists for {invoice_id}, will only download JSON")
    
    max_main_attempts = 2
    
    for main_attempt in range(max_main_attempts):
        try:
            if main_attempt > 0:
                logger.info(f"🔄 Main retry {main_attempt + 1} for {invoice_id}")
                driver.get(driver.current_url)
                wait_for_page_load()
                time.sleep(1.5)
            
            # Check status (cancelled/rejected/valid)
            invoice_status = check_if_cancelled()
            
            if invoice_status in ["Cancelled", "Rejected"]:
                logger.info(f"📥 Downloading {invoice_status} invoice: {invoice_id}")
            
            # Get issuer name and submission date (6 attempts)
            issuer_name, submission_date = get_issuer_name(max_attempts=6)
            logger.info(f"Processing {invoice_id} | Issuer: {issuer_name} | Submission: {submission_date} | Status: {invoice_status or 'Valid'}")
            
            # Download JSON only if not exists
            if json_exists:
                json_success = True
                logger.info(f"⏭ Skipping JSON download (already exists)")
            else:
                json_success = download_file('JSON', invoice_id, max_retries=5)
            time.sleep(1)
            
            # Download PDF only if not exists
            if pdf_exists:
                pdf_success = True
                logger.info(f"⏭ Skipping PDF download (already exists)")
            else:
                pdf_success = download_file('PDF', invoice_id, max_retries=5)
            
            # Evaluate results
            if json_success and pdf_success:
                status_label = f"{invoice_status}" if invoice_status else "Downloaded"
                logger.info(f"✓✓ Both files downloaded for {invoice_id} ({status_label})")
                add_invoice_to_excel(invoice_id, issuer_name, submission_date, status_label, taxpayer_name)
                if invoice_status:
                    scraping_summary['cancelled_invoices'] += 1
                return True
            elif json_success or pdf_success:
                file_type = "JSON" if json_success else "PDF"
                logger.warning(f"⚠ Only {file_type} downloaded for {invoice_id}")
                
                if main_attempt < max_main_attempts - 1:
                    logger.info("Will retry main download process")
                    continue
                
                status_label = f"Partial ({invoice_status})" if invoice_status else "Partial Download"
                add_invoice_to_excel(invoice_id, issuer_name, submission_date, status_label, taxpayer_name)
                return True
            else:
                logger.error(f"✗ Both downloads failed for {invoice_id}")
                
                if main_attempt < max_main_attempts - 1:
                    logger.info("Will retry main download process")
                    continue
                
                add_invoice_to_excel(invoice_id, issuer_name, submission_date, "Failed", taxpayer_name)
                return False
            
        except Exception as e:
            logger.error(f"Error in download_both_files: {e}")
            if main_attempt < max_main_attempts - 1:
                continue
            
            try:
                issuer_name, submission_date = get_issuer_name(max_attempts=3)
            except:
                issuer_name = "خطأ"
                submission_date = "خطأ"
            add_invoice_to_excel(invoice_id, issuer_name, submission_date, "Error", taxpayer_name)
            return False
    
    return False

def is_exact_date_match(invoice_date_str, target_date):
    """Check exact date match - CRITICAL"""
    try:
        if not invoice_date_str or not invoice_date_str.strip():
            return False, None
        
        invoice_date_str = invoice_date_str.strip()
        
        if "\n" in invoice_date_str:
            invoice_date_str = invoice_date_str.split("\n")[0]
        
        if " " in invoice_date_str:
            invoice_date_str = invoice_date_str.split(" ")[0]
        
        invoice_date_str = invoice_date_str.strip()
        
        invoice_date_obj = datetime.strptime(invoice_date_str, "%d/%m/%Y").date()
        is_match = invoice_date_obj == target_date
        
        return is_match, invoice_date_obj
        
    except:
        return False, None

def get_invoice_url_from_row(row):
    """Extract invoice URL from row"""
    try:
        link_elements = row.find_elements(By.CSS_SELECTOR, "a.griCellTitle")
        if not link_elements:
            return None, None
        
        invoice_id = link_elements[0].text.strip()
        invoice_url = link_elements[0].get_attribute('href')
        
        if not invoice_id or not invoice_url:
            return None, None
        
        return invoice_id, invoice_url
        
    except:
        return None, None

def login_and_select_taxpayer(taxpayer_name):
    """Login and select taxpayer"""
    try:
        logger.info(f"Logging in as: {taxpayer_name}")
        create_taxpayer_directories(taxpayer_name)
        
        driver.get("https://invoicing.eta.gov.eg/")
        wait_for_page_load()
        time.sleep(1)

        if not safe_send_keys((By.ID, "email"), EMAIL, timeout=12):
            raise Exception("Failed to enter email")
        
        if not safe_send_keys((By.ID, "Password"), PASSWORD, timeout=12):
            raise Exception("Failed to enter password")
        
        if not safe_click((By.XPATH, "//button[contains(text(), 'Login')]"), timeout=12):
            raise Exception("Failed to click login")
        
        wait_for_page_load()
        time.sleep(2)

        try:
            WebDriverWait(driver, 15).until(
                lambda d: len(d.find_elements(By.CSS_SELECTOR, "#SelectedTaxpayer option")) > 1
            )
            time.sleep(0.5)
        except:
            raise Exception("Dropdown not loaded")

        select_box = WebDriverWait(driver, 12).until(
            EC.presence_of_element_located((By.ID, "SelectedTaxpayer"))
        )
        select = Select(select_box)

        if taxpayer_name not in TAXPAYERS:
            raise Exception(f"Taxpayer not found: {taxpayer_name}")

        taxpayer_value = TAXPAYERS[taxpayer_name]
        select.select_by_value(taxpayer_value)
        time.sleep(0.6)

        if not safe_click((By.XPATH, "//button[contains(text(),'Select')]"), timeout=12):
            raise Exception("Failed to click select")
        
        wait_for_page_load()
        time.sleep(1.5)
        logger.info(f"✓ Logged in: {taxpayer_name}")
        
        scraping_summary['taxpayers_processed'].append({
            'name': taxpayer_name,
            'status': 'login_success',
            'invoices_downloaded': 0
        })
        
        return True
        
    except Exception as e:
        error_msg = f"Login failed for {taxpayer_name}: {e}"
        logger.error(error_msg)
        scraping_summary['errors'].append(error_msg)
        scraping_summary['taxpayers_processed'].append({
            'name': taxpayer_name,
            'status': 'login_failed',
            'error': str(e)
        })
        return False

def download_invoices_for_today(taxpayer_name):
    """Download invoices for target date - CRITICAL FUNCTION"""
    downloaded_invoice_ids = set()
    invoice_urls_to_process = []
    
    try:
        if not safe_click((By.ID, "invoices"), timeout=12):
            raise Exception("Failed to click invoices menu")
        
        wait_for_page_load()
        time.sleep(1)
        
        if not safe_click((By.ID, "recentDocuments"), timeout=12):
            raise Exception("Failed to click recent documents")
        
        wait_for_page_load()
        time.sleep(1.5)
        
        wait_for_rows_to_load()
        
        logger.info(f"Target date: {TARGET_DATE.strftime('%d/%m/%Y')}")

        page_number = 1
        total_invoices_found = 0
        target_date_invoices = 0
        consecutive_empty_pages = 0
        max_empty_pages = 2
        
        # PHASE 1: Collect all invoice URLs - CRITICAL
        logger.info("=" * 50)
        logger.info("PHASE 1: COLLECTING INVOICE URLs")
        logger.info("=" * 50)
        
        while True:
            try:
                logger.info(f"Scanning page {page_number}...")
                
                if not wait_for_rows_to_load(timeout=18):
                    logger.warning(f"Rows loading issue on page {page_number}")
                    consecutive_empty_pages += 1
                    if consecutive_empty_pages >= max_empty_pages:
                        logger.warning(f"Stopped after {consecutive_empty_pages} problematic pages")
                        break
                    
                    if not click_next_page():
                        break
                    page_number += 1
                    continue
                
                rows = driver.find_elements(By.CSS_SELECTOR, "div[role='row']")
                logger.info(f"Found {len(rows)} rows on page {page_number}")
                
                if len(rows) <= 1:
                    consecutive_empty_pages += 1
                    if consecutive_empty_pages >= max_empty_pages:
                        break
                    
                    if not click_next_page():
                        break
                    page_number += 1
                    continue

                page_target_invoices = 0
                should_stop = False
                older_dates_count = 0
                
                for i in range(len(rows)):
                    try:
                        rows = driver.find_elements(By.CSS_SELECTOR, "div[role='row']")
                        if i >= len(rows):
                            break
                        
                        row = rows[i]
                        date_cells = row.find_elements(By.CSS_SELECTOR, "div[data-automation-key='dateTimeReceived']")
                        
                        if not date_cells:
                            continue

                        invoice_date_str = date_cells[0].text.strip()
                        
                        if not invoice_date_str:
                            continue
                        
                        total_invoices_found += 1
                        
                        is_match, invoice_date_obj = is_exact_date_match(invoice_date_str, TARGET_DATE)
                        
                        if is_match:
                            invoice_id, invoice_url = get_invoice_url_from_row(row)
                            
                            if invoice_id and invoice_url:
                                if invoice_id in downloaded_invoice_ids:
                                    continue
                                
                                already_queued = any(inv['id'] == invoice_id for inv in invoice_urls_to_process)
                                if already_queued:
                                    continue
                                
                                logger.info(f"✓ Target invoice found: {invoice_id}")
                                invoice_urls_to_process.append({
                                    'id': invoice_id,
                                    'url': invoice_url,
                                    'date': invoice_date_str
                                })
                                target_date_invoices += 1
                                page_target_invoices += 1
                                consecutive_empty_pages = 0

                        elif invoice_date_obj and invoice_date_obj < TARGET_DATE:
                            older_dates_count += 1
                            
                            if older_dates_count >= 10 and target_date_invoices > 0:
                                logger.info(f"Found {older_dates_count} older invoices. Stopping pagination.")
                                should_stop = True
                                break

                    except StaleElementReferenceException:
                        continue
                    except Exception as e:
                        continue
                
                logger.info(f"Page {page_number} summary: {page_target_invoices} target invoices")
                
                if page_target_invoices > 0:
                    consecutive_empty_pages = 0
                
                if should_stop:
                    break

                if click_next_page():
                    page_number += 1
                    time.sleep(1)
                else:
                    logger.info("No more pages")
                    break

            except Exception as e:
                logger.error(f"Error on page {page_number}: {e}")
                scraping_summary['errors'].append(f"Page {page_number} error: {e}")
                
                try:
                    if click_next_page():
                        page_number += 1
                        continue
                except:
                    pass
                break

        logger.info("=" * 50)
        logger.info(f"PHASE 1 COMPLETE: {len(invoice_urls_to_process)} invoices collected")
        logger.info("=" * 50)
        
        # PHASE 2: Download all collected invoices - CRITICAL
        logger.info("=" * 50)
        logger.info("PHASE 2: DOWNLOADING INVOICES")
        logger.info("=" * 50)
        
        if len(invoice_urls_to_process) == 0:
            logger.warning(f"⚠️ No invoices found for {taxpayer_name} on {TARGET_DATE.strftime('%d/%m/%Y')}")
            logger.info(f"Total invoices scanned: {total_invoices_found}")
        else:
            logger.info(f"Starting downloads for {len(invoice_urls_to_process)} invoices...")
        
        for idx, invoice_info in enumerate(invoice_urls_to_process, 1):
            try:
                invoice_id = invoice_info['id']
                invoice_url = invoice_info['url']
                
                # Check if files already exist on disk
                json_file_path = os.path.join(current_json_dir, f"{invoice_id}.json")
                pdf_file_path = os.path.join(current_pdf_dir, f"{invoice_id}.pdf")
                
                json_exists = os.path.exists(json_file_path)
                pdf_exists = os.path.exists(pdf_file_path)
                
                if json_exists and pdf_exists:
                    logger.info(f"[{idx}/{len(invoice_urls_to_process)}] {invoice_id} already downloaded (both files exist)")
                    scraping_summary['total_invoices_downloaded'] += 1
                    downloaded_invoice_ids.add(invoice_id)
                    continue
                
                if invoice_id in downloaded_invoice_ids:
                    logger.info(f"[{idx}/{len(invoice_urls_to_process)}] {invoice_id} already processed in this session")
                    continue
                
                logger.info(f"[{idx}/{len(invoice_urls_to_process)}] Processing {invoice_id}")
                
                # Navigate to invoice
                driver.get(invoice_url)
                wait_for_page_load()
                time.sleep(1)
                
                if download_both_files(invoice_id, taxpayer_name):
                    scraping_summary['total_invoices_downloaded'] += 1
                    downloaded_invoice_ids.add(invoice_id)
                    logger.info(f"✓✓ Successfully completed {invoice_id}")
                else:
                    logger.warning(f"⚠ Failed to download {invoice_id}")
                
                time.sleep(0.5)

            except Exception as e:
                logger.error(f"Error processing {invoice_id}: {e}")
                scraping_summary['errors'].append(f"Failed {invoice_id}: {e}")
                continue

        logger.info("=" * 50)
        logger.info(f"PHASE 2 COMPLETE for {taxpayer_name}")
        logger.info(f"Total scanned: {total_invoices_found}")
        logger.info(f"Target found: {target_date_invoices}")
        logger.info(f"Downloaded: {len(downloaded_invoice_ids)}")
        logger.info(f"Downloaded IDs: {sorted(downloaded_invoice_ids)}")
        logger.info("=" * 50)

        for taxpayer in scraping_summary['taxpayers_processed']:
            if taxpayer['name'] == taxpayer_name:
                taxpayer['invoices_downloaded'] = len(downloaded_invoice_ids)
                break
                
    except Exception as e:
        error_msg = f"Error in download_invoices_for_today: {e}"
        logger.error(error_msg)
        scraping_summary['errors'].append(error_msg)

def logout():
    """Logout from the system"""
    try:
        wait_overlay_to_disappear()
        time.sleep(0.5)
        
        if safe_click((By.CSS_SELECTOR, "div.header-user-menu"), timeout=10):
            time.sleep(1)
            wait_overlay_to_disappear()
            
            if safe_click((By.XPATH, "//a[@href='/logout']"), timeout=10):
                time.sleep(1)
                
                try:
                    redirect_link = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a.PostLogoutRedirectUri"))
                    )
                    redirect_link.click()
                    logger.info("✓ Logged out")
                    wait_for_page_load()
                except:
                    logger.warning("Logout redirect timeout")
                    
    except Exception as e:
        logger.warning(f"Logout error: {e}")

def get_daily_totals():
    """Get total files downloaded today"""
    try:
        json_daily_dir = os.path.join(json_root, TARGET_DATE_STR)
        pdf_daily_dir = os.path.join(pdf_root, TARGET_DATE_STR)
        
        total_json = 0
        total_pdf = 0
        
        if os.path.exists(json_daily_dir):
            for root, dirs, files in os.walk(json_daily_dir):
                total_json += len([f for f in files if f.endswith('.json')])
        
        if os.path.exists(pdf_daily_dir):
            for root, dirs, files in os.walk(pdf_daily_dir):
                total_pdf += len([f for f in files if f.endswith('.pdf')])
        
        return total_json, total_pdf, total_json + total_pdf
        
    except:
        return 0, 0, 0

def verify_complete_download():
    """Verify all invoices downloaded completely"""
    try:
        logger.info("Verifying downloads...")
        
        verification_results = {
            'total_target_invoices': 0,
            'complete_downloads': 0,
            'partial_downloads': 0,
            'missing_downloads': 0,
            'missing_files': []
        }
        
        for taxpayer_name in TAXPAYERS.keys():
            taxpayer_json_dir = os.path.join(json_date_dir, taxpayer_name)
            taxpayer_pdf_dir = os.path.join(pdf_date_dir, taxpayer_name)
            
            if os.path.exists(taxpayer_json_dir) or os.path.exists(taxpayer_pdf_dir):
                json_files = []
                pdf_files = []
                
                if os.path.exists(taxpayer_json_dir):
                    json_files = [f for f in os.listdir(taxpayer_json_dir) if f.endswith('.json')]
                
                if os.path.exists(taxpayer_pdf_dir):
                    pdf_files = [f for f in os.listdir(taxpayer_pdf_dir) if f.endswith('.pdf')]
                
                json_ids = set([f.replace('.json', '') for f in json_files])
                pdf_ids = set([f.replace('.pdf', '') for f in pdf_files])
                
                all_ids = json_ids.union(pdf_ids)
                verification_results['total_target_invoices'] += len(all_ids)
                
                for invoice_id in all_ids:
                    has_json = invoice_id in json_ids
                    has_pdf = invoice_id in pdf_ids
                    
                    if has_json and has_pdf:
                        verification_results['complete_downloads'] += 1
                    elif has_json or has_pdf:
                        verification_results['partial_downloads'] += 1
                        missing_type = "PDF" if has_json else "JSON"
                        verification_results['missing_files'].append({
                            'taxpayer': taxpayer_name,
                            'invoice_id': invoice_id,
                            'missing_type': missing_type
                        })
                    else:
                        verification_results['missing_downloads'] += 1
        
        logger.info("=" * 60)
        logger.info("VERIFICATION RESULTS:")
        logger.info(f"  Total target invoices: {verification_results['total_target_invoices']}")
        logger.info(f"  Complete (JSON + PDF): {verification_results['complete_downloads']}")
        logger.info(f"  Partial downloads: {verification_results['partial_downloads']}")
        logger.info(f"  Missing downloads: {verification_results['missing_downloads']}")
        
        if verification_results['missing_files']:
            logger.warning("Missing files details:")
            for missing in verification_results['missing_files']:
                logger.warning(f"  - {missing['taxpayer']}: {missing['invoice_id']} missing {missing['missing_type']}")
        
        logger.info("=" * 60)
        
        return verification_results
        
    except Exception as e:
        logger.error(f"Verification error: {e}")
        return None

def main():
    """Main execution function"""
    scraping_summary['start_time'] = datetime.now()
    
    setup_excel_file()
    logger.info("=" * 60)
    logger.info("INVOICE SCRAPING PROCESS STARTED")
    logger.info("=" * 60)
    logger.info(f"Target date: {TARGET_DATE.strftime('%d/%m/%Y')}")
    logger.info(f"JSON directory: {json_root}/{TARGET_DATE_STR}/[taxpayer]/")
    logger.info(f"PDF directory: {pdf_root}/{TARGET_DATE_STR}/[taxpayer]/")
    logger.info("=" * 60)
    
    try:
        for taxpayer in TAXPAYERS.keys():
            logger.info(f"\n{'#' * 60}")
            logger.info(f"PROCESSING TAXPAYER: {taxpayer}")
            logger.info(f"{'#' * 60}\n")
            
            if login_and_select_taxpayer(taxpayer):
                download_invoices_for_today(taxpayer)
                logout()
                time.sleep(1.5)
            else:
                logger.error(f"Skipping {taxpayer} - login failed")
            
            time.sleep(1)
        
    except Exception as e:
        logger.error(f"Critical error in main: {e}")
        scraping_summary['errors'].append(f"Critical: {e}")
    
    finally:
        scraping_summary['end_time'] = datetime.now()
        driver.quit()
        
        # Save to Excel
        save_invoices_to_excel()
        
        # Verify downloads
        verification_results = verify_complete_download()
        
        # Calculate stats
        duration = scraping_summary['end_time'] - scraping_summary['start_time']
        daily_json, daily_pdf, daily_total = get_daily_totals()
        
        # Final summary
        logger.info("\n" + "=" * 60)
        logger.info("SCRAPING RUN COMPLETED")
        logger.info("=" * 60)
        logger.info(f"Run duration: {duration}")
        logger.info(f"This run - JSON: {scraping_summary['successful_downloads']['json']}, PDF: {scraping_summary['successful_downloads']['pdf']}")
        logger.info(f"Total invoices downloaded: {scraping_summary['total_invoices_downloaded']}")
        logger.info(f"Cancelled invoices skipped: {scraping_summary['cancelled_invoices']}")
        logger.info(f"Excel entries added: {len(scraping_summary['invoices_data'])}")
        logger.info(f"Errors encountered: {len(scraping_summary['errors'])}")
        
        if scraping_summary['errors']:
            logger.info("\nErrors details:")
            for i, error in enumerate(scraping_summary['errors'][:15], 1):
                logger.info(f"  {i}. {error}")
            if len(scraping_summary['errors']) > 15:
                logger.info(f"  ... and {len(scraping_summary['errors']) - 15} more errors")
        
        logger.info("\n" + "-" * 60)
        logger.info("DAILY TOTALS (all runs today):")
        logger.info(f"  Total JSON files: {daily_json}")
        logger.info(f"  Total PDF files: {daily_pdf}")
        logger.info(f"  Total files: {daily_total}")
        
        if verification_results:
            logger.info("\n" + "-" * 60)
            logger.info("DOWNLOAD COMPLETENESS:")
            logger.info(f"  Complete downloads: {verification_results['complete_downloads']}")
            logger.info(f"  Partial downloads: {verification_results['partial_downloads']}")
            logger.info(f"  Missing downloads: {verification_results['missing_downloads']}")
        
        logger.info("=" * 60)
        logger.info("END OF SCRAPING RUN")
        logger.info("=" * 60 + "\n")

if __name__ == "__main__":
    main()